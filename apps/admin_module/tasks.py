# apps/admin_module/tasks.py
import os
import logging
import tempfile
import zipfile
import json
from datetime import datetime, timedelta
from celery import shared_task
from django.utils import timezone
from django.conf import settings
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django_tenants.utils import tenant_context
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import io

from .models import DataExportLog

logger = logging.getLogger(__name__)

@shared_task(bind=True, max_retries=3)
def process_data_export(self, export_id):
    """
    Celery task to process data export requests.
    """
    try:
        # Get the export record
        export = DataExportLog.objects.get(id=export_id)
        
        # Mark as processing
        export.mark_as_processing()
        
        # Switch to tenant context
        with tenant_context(export.organization):
            # Process the export based on type
            if export.export_type == 'full_organization':
                file_path, file_size, records_count, tables_count = export_full_organization_data(export)
            elif export.export_type == 'audit_data':
                file_path, file_size, records_count, tables_count = export_audit_data(export)
            elif export.export_type == 'risk_data':
                file_path, file_size, records_count, tables_count = export_risk_data(export)
            elif export.export_type == 'compliance_data':
                file_path, file_size, records_count, tables_count = export_compliance_data(export)
            elif export.export_type == 'contracts_data':
                file_path, file_size, records_count, tables_count = export_contracts_data(export)
            elif export.export_type == 'legal_data':
                file_path, file_size, records_count, tables_count = export_legal_data(export)
            elif export.export_type == 'document_data':
                file_path, file_size, records_count, tables_count = export_document_data(export)
            elif export.export_type == 'user_data':
                file_path, file_size, records_count, tables_count = export_user_data(export)
            elif export.export_type == 'custom':
                file_path, file_size, records_count, tables_count = export_custom_data(export)
            else:
                raise ValueError(f"Unknown export type: {export.export_type}")
            
            # Mark as completed
            export.mark_as_completed(file_path, file_size, records_count, tables_count)
            
            # Send notification
            send_export_completion_notification.delay(export_id)
            
            logger.info(f"Export {export_id} completed successfully")
            
    except DataExportLog.DoesNotExist:
        logger.error(f"Export {export_id} not found")
    except Exception as exc:
        logger.error(f"Export {export_id} failed: {exc}")
        
        # Mark as failed
        try:
            export = DataExportLog.objects.get(id=export_id)
            export.mark_as_failed(str(exc))
        except DataExportLog.DoesNotExist:
            pass
        
        # Retry if possible
        if self.request.retries < self.max_retries:
            raise self.retry(countdown=60 * (2 ** self.request.retries))
        else:
            logger.error(f"Export {export_id} failed after {self.max_retries} retries")

def export_full_organization_data(export):
    """Export all organization data."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    records_count = 0
    tables_count = 0
    
    # Export each data type
    data_types = [
        ('audit_workplans', 'Audit Workplans'),
        ('audit_engagements', 'Audit Engagements'),
        ('audit_issues', 'Audit Issues'),
        ('risk_assessments', 'Risk Assessments'),
        ('risk_controls', 'Risk Controls'),
        ('compliance_requirements', 'Compliance Requirements'),
        ('compliance_obligations', 'Compliance Obligations'),
        ('contracts', 'Contracts'),
        ('contract_milestones', 'Contract Milestones'),
        ('legal_cases', 'Legal Cases'),
        ('legal_tasks', 'Legal Tasks'),
        ('documents', 'Documents'),
    ]
    
    for data_type, sheet_name in data_types:
        try:
            sheet_data = get_data_for_type(data_type, export)
            if sheet_data:
                ws = wb.create_sheet(sheet_name)
                add_data_to_sheet(ws, sheet_data)
                records_count += len(sheet_data.get('data', []))
                tables_count += 1
        except Exception as e:
            logger.error(f"Error exporting {data_type}: {e}")
    
    # Save to file
    file_path = save_workbook(wb, export, 'full_organization')
    file_size = get_file_size(file_path)
    
    return file_path, file_size, records_count, tables_count

def export_audit_data(export):
    """Export audit-related data."""
    wb = Workbook()
    wb.remove(wb.active)
    
    records_count = 0
    tables_count = 0
    
    audit_types = [
        ('audit_workplans', 'Audit Workplans'),
        ('audit_engagements', 'Audit Engagements'),
        ('audit_issues', 'Audit Issues'),
    ]
    
    for data_type, sheet_name in audit_types:
        try:
            sheet_data = get_data_for_type(data_type, export)
            if sheet_data:
                ws = wb.create_sheet(sheet_name)
                add_data_to_sheet(ws, sheet_data)
                records_count += len(sheet_data.get('data', []))
                tables_count += 1
        except Exception as e:
            logger.error(f"Error exporting {data_type}: {e}")
    
    file_path = save_workbook(wb, export, 'audit_data')
    file_size = get_file_size(file_path)
    
    return file_path, file_size, records_count, tables_count

def export_risk_data(export):
    """Export risk-related data."""
    wb = Workbook()
    wb.remove(wb.active)
    
    records_count = 0
    tables_count = 0
    
    risk_types = [
        ('risk_assessments', 'Risk Assessments'),
        ('risk_controls', 'Risk Controls'),
    ]
    
    for data_type, sheet_name in risk_types:
        try:
            sheet_data = get_data_for_type(data_type, export)
            if sheet_data:
                ws = wb.create_sheet(sheet_name)
                add_data_to_sheet(ws, sheet_data)
                records_count += len(sheet_data.get('data', []))
                tables_count += 1
        except Exception as e:
            logger.error(f"Error exporting {data_type}: {e}")
    
    file_path = save_workbook(wb, export, 'risk_data')
    file_size = get_file_size(file_path)
    
    return file_path, file_size, records_count, tables_count

def export_compliance_data(export):
    """Export compliance-related data."""
    wb = Workbook()
    wb.remove(wb.active)
    
    records_count = 0
    tables_count = 0
    
    compliance_types = [
        ('compliance_requirements', 'Compliance Requirements'),
        ('compliance_obligations', 'Compliance Obligations'),
    ]
    
    for data_type, sheet_name in compliance_types:
        try:
            sheet_data = get_data_for_type(data_type, export)
            if sheet_data:
                ws = wb.create_sheet(sheet_name)
                add_data_to_sheet(ws, sheet_data)
                records_count += len(sheet_data.get('data', []))
                tables_count += 1
        except Exception as e:
            logger.error(f"Error exporting {data_type}: {e}")
    
    file_path = save_workbook(wb, export, 'compliance_data')
    file_size = get_file_size(file_path)
    
    return file_path, file_size, records_count, tables_count

def export_contracts_data(export):
    """Export contracts-related data."""
    wb = Workbook()
    wb.remove(wb.active)
    
    records_count = 0
    tables_count = 0
    
    contract_types = [
        ('contracts', 'Contracts'),
        ('contract_milestones', 'Contract Milestones'),
    ]
    
    for data_type, sheet_name in contract_types:
        try:
            sheet_data = get_data_for_type(data_type, export)
            if sheet_data:
                ws = wb.create_sheet(sheet_name)
                add_data_to_sheet(ws, sheet_data)
                records_count += len(sheet_data.get('data', []))
                tables_count += 1
        except Exception as e:
            logger.error(f"Error exporting {data_type}: {e}")
    
    file_path = save_workbook(wb, export, 'contracts_data')
    file_size = get_file_size(file_path)
    
    return file_path, file_size, records_count, tables_count

def export_legal_data(export):
    """Export legal-related data."""
    wb = Workbook()
    wb.remove(wb.active)
    
    records_count = 0
    tables_count = 0
    
    legal_types = [
        ('legal_cases', 'Legal Cases'),
        ('legal_tasks', 'Legal Tasks'),
    ]
    
    for data_type, sheet_name in legal_types:
        try:
            sheet_data = get_data_for_type(data_type, export)
            if sheet_data:
                ws = wb.create_sheet(sheet_name)
                add_data_to_sheet(ws, sheet_data)
                records_count += len(sheet_data.get('data', []))
                tables_count += 1
        except Exception as e:
            logger.error(f"Error exporting {data_type}: {e}")
    
    file_path = save_workbook(wb, export, 'legal_data')
    file_size = get_file_size(file_path)
    
    return file_path, file_size, records_count, tables_count

def export_document_data(export):
    """Export document management data."""
    wb = Workbook()
    wb.remove(wb.active)
    
    records_count = 0
    tables_count = 0
    
    try:
        sheet_data = get_data_for_type('documents', export)
        if sheet_data:
            ws = wb.create_sheet('Documents')
            add_data_to_sheet(ws, sheet_data)
            records_count += len(sheet_data.get('data', []))
            tables_count += 1
    except Exception as e:
        logger.error(f"Error exporting documents: {e}")
    
    file_path = save_workbook(wb, export, 'document_data')
    file_size = get_file_size(file_path)
    
    return file_path, file_size, records_count, tables_count

def export_user_data(export):
    """Export user-related data."""
    wb = Workbook()
    wb.remove(wb.active)
    
    records_count = 0
    tables_count = 0
    
    try:
        sheet_data = get_data_for_type('users', export)
        if sheet_data:
            ws = wb.create_sheet('Users')
            add_data_to_sheet(ws, sheet_data)
            records_count += len(sheet_data.get('data', []))
            tables_count += 1
    except Exception as e:
        logger.error(f"Error exporting users: {e}")
    
    file_path = save_workbook(wb, export, 'user_data')
    file_size = get_file_size(file_path)
    
    return file_path, file_size, records_count, tables_count

def export_custom_data(export):
    """Export custom selected data."""
    wb = Workbook()
    wb.remove(wb.active)
    
    records_count = 0
    tables_count = 0
    
    custom_selection = export.custom_selection
    selected_models = custom_selection.get('models', [])
    
    for model_type in selected_models:
        try:
            sheet_data = get_data_for_type(model_type, export)
            if sheet_data:
                sheet_name = get_sheet_name_for_type(model_type)
                ws = wb.create_sheet(sheet_name)
                add_data_to_sheet(ws, sheet_data)
                records_count += len(sheet_data.get('data', []))
                tables_count += 1
        except Exception as e:
            logger.error(f"Error exporting {model_type}: {e}")
    
    file_path = save_workbook(wb, export, 'custom_data')
    file_size = get_file_size(file_path)
    
    return file_path, file_size, records_count, tables_count

def get_data_for_type(data_type, export):
    """Get data for a specific type."""
    # Apply date filters
    date_from = export.custom_selection.get('date_from')
    date_to = export.custom_selection.get('date_to')
    
    # Import models dynamically
    if data_type == 'audit_workplans':
        from audit.models import AuditWorkplan
        queryset = AuditWorkplan.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'title', 'description', 'status', 'created_at', 'updated_at'])
    
    elif data_type == 'audit_engagements':
        from audit.models import Engagement
        queryset = Engagement.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'name', 'description', 'status', 'start_date', 'end_date', 'created_at'])
    
    elif data_type == 'audit_issues':
        from audit.models import Issue
        queryset = Issue.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'title', 'description', 'severity', 'status', 'created_at', 'updated_at'])
    
    elif data_type == 'risk_assessments':
        from risk.models import RiskAssessment
        queryset = RiskAssessment.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'title', 'description', 'risk_level', 'status', 'created_at'])
    
    elif data_type == 'risk_controls':
        from risk.models import Control
        queryset = Control.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'name', 'description', 'control_type', 'effectiveness', 'created_at'])
    
    elif data_type == 'compliance_requirements':
        from compliance.models import ComplianceRequirement
        queryset = ComplianceRequirement.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'title', 'description', 'category', 'status', 'created_at'])
    
    elif data_type == 'compliance_obligations':
        from compliance.models import ComplianceObligation
        queryset = ComplianceObligation.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'title', 'description', 'due_date', 'status', 'created_at'])
    
    elif data_type == 'contracts':
        from contracts.models import Contract
        queryset = Contract.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'title', 'contract_number', 'value', 'start_date', 'end_date', 'status', 'created_at'])
    
    elif data_type == 'contract_milestones':
        from contracts.models import Milestone
        queryset = Milestone.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'title', 'description', 'due_date', 'status', 'created_at'])
    
    elif data_type == 'legal_cases':
        from legal.models import LegalCase
        queryset = LegalCase.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'title', 'case_number', 'status', 'priority', 'created_at'])
    
    elif data_type == 'legal_tasks':
        from legal.models import LegalTask
        queryset = LegalTask.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'title', 'description', 'due_date', 'status', 'priority', 'created_at'])
    
    elif data_type == 'documents':
        from document_management.models import Document
        queryset = Document.objects.all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'title', 'document_type', 'file_path', 'created_at', 'updated_at'])
    
    elif data_type == 'users':
        from users.models import CustomUser
        queryset = CustomUser.objects.filter(organization=export.organization)
        if date_from:
            queryset = queryset.filter(date_joined__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date_joined__date__lte=date_to)
        return serialize_queryset(queryset, ['id', 'email', 'first_name', 'last_name', 'role', 'is_active', 'date_joined'])
    
    return None

def serialize_queryset(queryset, fields):
    """Serialize a queryset to a list of dictionaries."""
    data = []
    for obj in queryset:
        row = {}
        for field in fields:
            value = getattr(obj, field, None)
            if hasattr(value, 'strftime'):  # Handle datetime fields
                value = value.strftime('%Y-%m-%d %H:%M:%S') if value else None
            row[field] = value
        data.append(row)
    
    return {
        'headers': fields,
        'data': data
    }

def add_data_to_sheet(worksheet, sheet_data):
    """Add data to an Excel worksheet."""
    headers = sheet_data['headers']
    data = sheet_data['data']
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row, record in enumerate(data, 2):
        for col, header in enumerate(headers, 1):
            value = record.get(header, '')
            worksheet.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def save_workbook(workbook, export, prefix):
    """Save workbook to file storage."""
    # Create temporary file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        workbook.save(tmp_file.name)
        tmp_file_path = tmp_file.name
    
    # Generate filename
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f"data_export_{prefix}_{export.organization.customer_code}_{timestamp}.xlsx"
    
    # Save to storage
    file_path = f"data_exports/{export.organization.customer_code}/{filename}"
    
    with open(tmp_file_path, 'rb') as f:
        default_storage.save(file_path, ContentFile(f.read()))
    
    # Clean up temporary file
    os.unlink(tmp_file_path)
    
    return file_path

def get_file_size(file_path):
    """Get file size in bytes."""
    try:
        return default_storage.size(file_path)
    except:
        return 0

def get_sheet_name_for_type(data_type):
    """Get sheet name for data type."""
    sheet_names = {
        'audit_workplans': 'Audit Workplans',
        'audit_engagements': 'Audit Engagements',
        'audit_issues': 'Audit Issues',
        'risk_assessments': 'Risk Assessments',
        'risk_controls': 'Risk Controls',
        'compliance_requirements': 'Compliance Requirements',
        'compliance_obligations': 'Compliance Obligations',
        'contracts': 'Contracts',
        'contract_milestones': 'Contract Milestones',
        'legal_cases': 'Legal Cases',
        'legal_tasks': 'Legal Tasks',
        'documents': 'Documents',
        'users': 'Users',
    }
    return sheet_names.get(data_type, data_type.title().replace('_', ' '))

@shared_task
def send_export_completion_notification(export_id):
    """Send notification when export is completed."""
    try:
        export = DataExportLog.objects.get(id=export_id)
        
        # Send email notification
        from django.core.mail import send_mail
        from django.template.loader import render_to_string
        
        subject = f"Data Export Completed - {export.export_type}"
        context = {
            'export': export,
            'user': export.requested_by,
            'organization': export.organization,
        }
        
        html_message = render_to_string('admin_module/email/export_completed.html', context)
        text_message = render_to_string('admin_module/email/export_completed.txt', context)
        
        send_mail(
            subject=subject,
            message=text_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[export.requested_by.email],
            html_message=html_message,
            fail_silently=False,
        )
        
        logger.info(f"Export completion notification sent for export {export_id}")
        
    except DataExportLog.DoesNotExist:
        logger.error(f"Export {export_id} not found for notification")
    except Exception as e:
        logger.error(f"Failed to send export completion notification for {export_id}: {e}")

@shared_task
def cleanup_expired_exports():
    """Clean up expired export files."""
    try:
        expired_exports = DataExportLog.objects.filter(
            file_expires_at__lt=timezone.now(),
            is_deleted=False
        )
        
        deleted_count = 0
        for export in expired_exports:
            try:
                # Delete file if it exists
                if export.file_path and default_storage.exists(export.file_path):
                    default_storage.delete(export.file_path)
                
                # Mark as deleted
                export.is_deleted = True
                export.save(update_fields=['is_deleted'])
                deleted_count += 1
                
            except Exception as e:
                logger.error(f"Error cleaning up export {export.id}: {e}")
        
        logger.info(f"Cleanup completed. {deleted_count} expired files deleted.")
        
    except Exception as e:
        logger.error(f"Error during cleanup: {e}")
