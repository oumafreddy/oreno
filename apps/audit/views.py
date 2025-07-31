# apps/audit/views.py
from django.views.decorators.http import require_http_methods
from django.urls import reverse, reverse_lazy
from django.core.exceptions import ValidationError
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from . import models
from users.models import CustomUser
from .models.workplan import AuditWorkplan


# Utility functions to avoid AttributeError with htmx
def is_htmx_request(request):
    """Safely check if a request is an HTMX request without AttributeError"""
    return request.headers and request.headers.get('HX-Request') == 'true'

# Utility function to safely retrieve issue_pk from various sources
def safe_get_issue_pk(view_instance, raise_error=False):
    """Safely retrieve issue_pk from various sources including URL params, GET/POST data, or view object.
    
    Args:
        view_instance: The view instance (needs to have request and kwargs)
        raise_error: Whether to raise ValueError if issue_pk not found (default: False)
        
    Returns:
        The issue_pk value or None if not found and raise_error is False
    """
    # Try from URL parameters or request data
    issue_pk = view_instance.kwargs.get("issue_pk") or \
               view_instance.request.GET.get("issue_pk") or \
               view_instance.request.POST.get("issue_pk")
    
    # If not found, try to get from the view object
    if not issue_pk and hasattr(view_instance, 'object') and view_instance.object:
        if hasattr(view_instance.object, 'issue_id') and view_instance.object.issue_id:
            return view_instance.object.issue_id
    
    # If still not found, try to get the object from the database
    if not issue_pk and hasattr(view_instance, 'get_object') and hasattr(view_instance, 'kwargs'):
        try:
            if hasattr(view_instance.kwargs, 'pk') or 'pk' in view_instance.kwargs:
                obj = view_instance.get_object()
                if hasattr(obj, 'issue_id') and obj.issue_id:
                    return obj.issue_id
        except Exception:
            # Silently handle any errors during object retrieval
            pass
    
    # If we need to raise an error and we don't have an issue_pk
    if not issue_pk and raise_error:
        raise ValueError("issue_pk is required")
        
    return issue_pk

# Add htmx attribute to request objects to ensure backward compatibility
class RequestWrapper:
    def __init__(self, get_response):
        self.get_response = get_response
        
    def __call__(self, request):
        # Add htmx attribute if it doesn't exist
        if not hasattr(request, 'htmx'):
            request.htmx = is_htmx_request(request)
        return self.get_response(request)

# Monkey patch the request object to ensure htmx attribute
import types
from django.http.request import HttpRequest
def _htmx_property(self):
    return is_htmx_request(self)
HttpRequest.htmx = property(_htmx_property)


from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse_lazy
from rest_framework import permissions
from django.conf import settings
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
)
from django.contrib.messages.views import SuccessMessageMixin
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import PermissionDenied
from django.db import transaction, DatabaseError, NotSupportedError
from django.db.models import Q, F, Case, When
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils.translation import gettext as _
from django.utils import timezone
from django.contrib.contenttypes.models import ContentType
from .models import Approval
from django.urls import reverse_lazy
from django.views.decorators.http import require_http_methods as require_HTTP_methods
import logging

logger = logging.getLogger(__name__)
from django.utils.translation import gettext_lazy as _
from django.contrib.auth.decorators import login_required, permission_required
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView,
    TemplateView
)
from django_fsm import can_proceed

from core.mixins import OrganizationMixin, OrganizationPermissionMixin
from core.decorators import skip_org_check
from organizations.models import Organization

from .models import AuditWorkplan, Engagement, Issue, Approval, Notification, IssueWorkingPaper, Note, FollowUpAction, IssueRetest, Objective, Procedure
from .forms import (
    AuditWorkplanForm, EngagementForm, IssueForm,
    ApprovalForm, WorkplanFilterForm, EngagementFilterForm, IssueFilterForm,
    ProcedureForm, FollowUpActionForm, IssueRetestForm, NoteForm,
    RecommendationForm, IssueWorkingPaperForm, ObjectiveForm, RiskForm
)

from rest_framework import generics, permissions, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from .serializers import (
    AuditWorkplanSerializer, EngagementSerializer, IssueSerializer,
    ApprovalSerializer, NoteSerializer, NotificationSerializer
)

from core.permissions import IsTenantMember
from django.contrib.contenttypes.models import ContentType
from users.permissions import IsOrgAdmin, IsOrgManagerOrReadOnly, HasOrgAdminAccess
from core.mixins.organization import OrganizationScopedQuerysetMixin
from .mixins import AuditOrganizationScopedMixin, OrganizationScopedApiMixin

import csv
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from django.utils.encoding import smart_str

from .models.objective import Objective

from .models.risk import Risk
from .models.engagement import Engagement
from .models.issue import Issue
from users.permissions import IsOrgManagerOrReadOnly

from .models.procedure import Procedure
from .models.followupaction import FollowUpAction
from .models.issueretest import IssueRetest
from .models.note import Note
from .models.recommendation import Recommendation

from django.http import JsonResponse
from django.template.loader import render_to_string
from django.db.models import Count
from django.utils import timezone
from django.views.decorators.http import require_GET

# ─── MIXINS ─────────────────────────────────────────────────────────────────
class AuditPermissionMixin(AuditOrganizationScopedMixin):
    """
    Verify that current user has audit permissions for the organization.
    Inherits from AuditOrganizationScopedMixin to ensure consistent organization scoping.
    """
    def dispatch(self, request, *args, **kwargs):
        # Base permission check - user must be authenticated
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        
        # Check if user has an active organization
        if not hasattr(request.user, 'active_organization') or not request.user.active_organization:
            messages.warning(request, _('You need to select an active organization'))
            return redirect('users:organization-list')
            
        # Check if user has the necessary permissions for the active organization
        # We could implement role-based checks here in the future
        
        return super().dispatch(request, *args, **kwargs)

# ─── RISK VIEWS ─────────────────────────────────────────────────────────────────
class RiskListView(AuditPermissionMixin, ListView):
    model = Risk
    template_name = 'audit/risk_list.html'
    context_object_name = 'risks'
    paginate_by = 20
    
    def get_queryset(self):
        # Base queryset already filtered by organization in AuditOrganizationScopedMixin
        queryset = super().get_queryset()
        
        # Apply filters based on GET parameters
        q = self.request.GET.get('q', '')
        if q:
            queryset = queryset.filter(
                Q(title__icontains=q) | Q(description__icontains=q)
            )
            
        # Filter by objective if specified
        objective_id = self.request.GET.get('objective_id') or self.kwargs.get('objective_id')
        if objective_id:
            queryset = queryset.filter(objective_id=objective_id)
            
        # Apply sorting
        queryset = queryset.order_by('objective__title', 'title')
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add objective to context if filtering by objective
        objective_id = self.request.GET.get('objective_id') or self.kwargs.get('objective_id')
        if objective_id:
            context['objective'] = get_object_or_404(
                Objective, 
                id=objective_id, 
                engagement__organization=self.request.organization
            )
            
        # Add filter parameters
        context['q'] = self.request.GET.get('q', '')
        
        return context


class RiskDetailView(AuditPermissionMixin, LoginRequiredMixin, DetailView):
    model = Risk
    template_name = 'audit/risk_detail.html'
    context_object_name = 'risk'
    
    def get_queryset(self):
        # Ensure user can only view risks in their organization
        return super().get_queryset().filter(organization=self.request.organization)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Use the correct related_name for procedures
        context['procedures'] = self.object.procedures.all().order_by('order')
        # Add the objective for breadcrumb navigation
        if self.object.objective:
            context['objective'] = self.object.objective
            context['engagement'] = self.object.objective.engagement
        return context


class RiskCreateView(AuditPermissionMixin, LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Risk
    form_class = RiskForm
    template_name = 'audit/risk_form.html'
    success_message = _('Risk "%(title)s" was created successfully')
    
    def get_template_names(self):
        try:
            # Use modal template for HTMX requests
            if is_htmx_request(self.request):
                return ['audit/risk_modal_form.html']
            return [self.template_name]
        except Exception as e:
            logger.error(f"Error in get_template_names: {str(e)}", exc_info=True)
            raise
    
    def get_success_url(self):
        try:
            # Return to objective detail if creating from an objective
            if self.object.objective:
                return reverse_lazy('audit:objective-detail', kwargs={'pk': self.object.objective.pk})
            # Default to engagement detail if creating from engagement
            elif hasattr(self.object, 'engagement_context') and self.object.engagement_context:
                return reverse_lazy('audit:engagement-detail', kwargs={'pk': self.object.engagement_context.pk})
            # Default to risk list
            return reverse_lazy('audit:risk-list')
        except Exception as e:
            logger.error(f"Error in get_success_url: {str(e)}", exc_info=True)
            raise
    
    def get_form_kwargs(self):
        try:
            # Add organization and context to form kwargs
            kwargs = super().get_form_kwargs()
            kwargs['organization'] = self.request.organization
            
            # Add objective_pk if available
            objective_id = self.request.GET.get('objective_id') or self.kwargs.get('objective_id')
            if objective_id:
                kwargs['objective_pk'] = objective_id
                
            # Add engagement_pk if available
            engagement_pk = self.kwargs.get('engagement_pk')
            if engagement_pk:
                kwargs['engagement_pk'] = engagement_pk
                
            return kwargs
        except Exception as e:
            logger.error(f"Error in get_form_kwargs: {str(e)}", exc_info=True)
            raise
    
    def form_valid(self, form):
        try:
            # Set organization and user fields
            form.instance.organization = self.request.organization
            form.instance.created_by = self.request.user
            form.instance.updated_by = self.request.user
            engagement_pk = self.kwargs.get('engagement_pk')
            if engagement_pk and not form.instance.objective:
                try:
                    engagement = Engagement.objects.get(pk=engagement_pk, organization=self.request.organization)
                    form.instance.engagement_context = engagement
                except Engagement.DoesNotExist:
                    logger.warning(f"Engagement with pk {engagement_pk} not found for organization {self.request.organization}")
                except Exception as e:
                    logger.error(f"Error setting engagement context: {str(e)}", exc_info=True)
            
            # Call super().form_valid() first to save the object and set self.object
            response = super().form_valid(form)
            
            # --- HTMX: Return HTML partial, not JSON ---
            if is_htmx_request(self.request):
                if hasattr(self.object, 'objective') and self.object.objective:
                    risks = Risk.objects.filter(objective=self.object.objective).order_by('order')
                    html_list = render_to_string('audit/_risk_list_partial.html', {
                        'risks': risks,
                        'objective': self.object.objective,
                    }, request=self.request)
                    return HttpResponse(html_list)
                else:
                    engagement_pk = self.kwargs.get('engagement_pk')
                    if engagement_pk:
                        redirect_url = reverse_lazy('audit:engagement-detail', kwargs={'pk': engagement_pk})
                    else:
                        redirect_url = self.get_success_url()
                    # Return a simple redirect script for HTMX
                    return HttpResponse(f'<script>window.location.href="{redirect_url}";</script>')
            
            # --- AJAX: Return JSON ---
            if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                if hasattr(self.object, 'objective') and self.object.objective:
                    risks = Risk.objects.filter(objective=self.object.objective).order_by('order')
                    html_list = render_to_string('audit/_risk_list_partial.html', {
                        'risks': risks,
                        'objective': self.object.objective,
                    }, request=self.request)
                    return JsonResponse({
                        'success': True,
                        'html_list': html_list,
                        'message': self.success_message % {'title': self.object.title}
                    })
                else:
                    engagement_pk = self.kwargs.get('engagement_pk')
                    if engagement_pk:
                        redirect_url = reverse_lazy('audit:engagement-detail', kwargs={'pk': engagement_pk})
                    else:
                        redirect_url = self.get_success_url()
                    return JsonResponse({
                        'success': True,
                        'redirect': str(redirect_url),
                        'message': self.success_message % {'title': self.object.title}
                    })
            return response
        except ValidationError as e:
            logger.error(f"Validation error in form_valid: {str(e)}", exc_info=True)
            return self.form_invalid(form)
        except DatabaseError as e:
            logger.error(f"Database error in form_valid: {str(e)}", exc_info=True)
            raise
        except Exception as e:
            logger.error(f"Unexpected error in form_valid: {str(e)}", exc_info=True)
            raise
    
    def form_invalid(self, form):
        try:
            if is_htmx_request(self.request):
                html_form = render_to_string(
                    'audit/risk_modal_form.html',
                    self.get_context_data(form=form),
                    request=self.request
                )
                return JsonResponse({'success': False, 'html_form': html_form})
            return super().form_invalid(form)
        except Exception as e:
            logger.error(f"Error in form_invalid: {str(e)}", exc_info=True)
            raise
    
    def get_context_data(self, **kwargs):
        try:
            context = super().get_context_data(**kwargs)
            
            # Add objective to context if creating from an objective
            objective_id = self.request.GET.get('objective_id') or self.kwargs.get('objective_id')
            if objective_id:
                try:
                    objective = Objective.objects.get(
                        id=objective_id, 
                        engagement__organization=self.request.organization
                    )
                    context['objective'] = objective
                    context['engagement'] = objective.engagement
                except Objective.DoesNotExist:
                    logger.warning(f"Objective with id {objective_id} not found for organization {self.request.organization}")
                except Exception as e:
                    logger.error(f"Error getting objective context: {str(e)}", exc_info=True)
            
            # Add engagement to context if creating from an engagement
            engagement_pk = self.kwargs.get('engagement_pk')
            if engagement_pk and 'engagement' not in context:
                try:
                    engagement = Engagement.objects.get(
                        pk=engagement_pk,
                        organization=self.request.organization
                    )
                    context['engagement'] = engagement
                except Engagement.DoesNotExist:
                    logger.warning(f"Engagement with pk {engagement_pk} not found for organization {self.request.organization}")
                except Exception as e:
                    logger.error(f"Error getting engagement context: {str(e)}", exc_info=True)
            
            return context
        except Exception as e:
            logger.error(f"Error in get_context_data: {str(e)}", exc_info=True)
            raise


class RiskUpdateView(AuditPermissionMixin, LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Risk
    form_class = RiskForm
    template_name = 'audit/risk_form.html'
    success_message = _('Risk "%(title)s" was updated successfully')
    
    def get_template_names(self):
        # Use modal template for HTMX requests
        if is_htmx_request(self.request):
            return ['audit/risk_modal_form.html']
        return [self.template_name]
    
    def get_queryset(self):
        # Ensure user can only update risks in their organization
        return super().get_queryset().filter(organization=self.request.organization)
    
    def get_success_url(self):
        # Return to the risk detail page
        return reverse_lazy('audit:risk-detail', kwargs={'pk': self.object.pk})
    
    def get_form_kwargs(self):
        # Add organization to form kwargs
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        # Filter objectives by the engagement of the current risk
        risk = self.get_object()
        if risk and risk.objective and risk.objective.engagement_id:
            kwargs['engagement_pk'] = risk.objective.engagement_id
        return kwargs
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        if is_htmx_request(self.request):
            if self.object.objective:
                risks = Risk.objects.filter(objective=self.object.objective).order_by('order')
                html = render_to_string('audit/_risk_list_partial.html', {
                    'risks': risks,
                    'objective': self.object.objective,
                }, request=self.request)
                return HttpResponse(html)
            else:
                return HttpResponse(f'<script>window.location.href="{self.get_success_url()}";</script>')
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            if self.object.objective:
                risks = Risk.objects.filter(objective=self.object.objective).order_by('order')
                html = render_to_string('audit/_risk_list_partial.html', {
                    'risks': risks,
                    'objective': self.object.objective,
                }, request=self.request)
                return JsonResponse({
                    'success': True,
                    'html': html,
                    'message': self.success_message % {'title': self.object.title}
                })
            else:
                return JsonResponse({
                    'success': True,
                    'redirect': self.get_success_url(),
                    'message': self.success_message % {'title': self.object.title}
                })
        return response
    
    def form_invalid(self, form):
        if is_htmx_request(self.request):
            html_form = render_to_string(
                'audit/risk_modal_form.html',
                self.get_context_data(form=form),
                request=self.request
            )
            return JsonResponse({'success': False, 'html_form': html_form})
        return super().form_invalid(form)


class RiskDeleteView(AuditPermissionMixin, LoginRequiredMixin, DeleteView):
    model = Risk
    template_name = 'audit/risk_confirm_delete.html'

    def get_template_names(self):
        # Use modal template for HTMX requests
        if is_htmx_request(self.request):
            return ['audit/risk_confirm_delete_modal.html']
        return [self.template_name]

    def get_queryset(self):
        # Ensure user can only delete risks in their organization
        return super().get_queryset().filter(organization=self.request.organization)

    def get_success_url(self):
        # Redirect to parent objective detail if available, else engagement, else risk list
        if self.object.objective_id:
            return reverse_lazy('audit:objective-detail', kwargs={'pk': self.object.objective_id})
        elif self.object.objective and self.object.objective.engagement_id:
            return reverse_lazy('audit:engagement-detail', kwargs={'pk': self.object.objective.engagement_id})
        return reverse_lazy('audit:risk-list')

from .models.issue_working_paper import IssueWorkingPaper
from .forms import IssueWorkingPaperForm
from .serializers import IssueWorkingPaperSerializer

# ─── MIXINS ──────────────────────────────────────────────────────────────────
# AuditPermissionMixin moved to the top of the file

# ─── WORKPLAN VIEWS ──────────────────────────────────────────────────────────
class WorkplanListView(AuditPermissionMixin, ListView):
    model = AuditWorkplan
    template_name = 'audit/workplan_list.html'
    context_object_name = 'workplans'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        organization = self.request.organization
        queryset = queryset.filter(organization=organization).select_related(
            'created_by', 'updated_by'
        ).prefetch_related('engagements')
        form = WorkplanFilterForm(self.request.GET)
        if form.is_valid():
            q = form.cleaned_data.get('q')
            if q:
                queryset = queryset.filter(Q(name__icontains=q) | Q(code__icontains=q))
            status = form.cleaned_data.get('status')
            if status:
                queryset = queryset.filter(approval_status=status)
            fiscal_year = form.cleaned_data.get('fiscal_year')
            if fiscal_year:
                queryset = queryset.filter(fiscal_year=fiscal_year)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = WorkplanFilterForm(self.request.GET)
        return context

class WorkplanDetailView(AuditPermissionMixin, DetailView):
    model = AuditWorkplan
    template_name = 'audit/workplan_detail.html'
    context_object_name = 'workplan'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        workplan = self.object
        
        # Add related data efficiently
        context.update({
            'engagements': workplan.engagements.all().select_related(
                'assigned_to', 'assigned_by'
            ),
            'approvals': workplan.approvals.all().select_related(
                'requester', 'approver'
            ),
            'can_submit': can_proceed(workplan.submit_for_approval),
            'can_approve': can_proceed(workplan.approve),
            'can_reject': can_proceed(workplan.reject),
        })
        return context

class WorkplanCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = AuditWorkplan
    form_class = AuditWorkplanForm
    template_name = 'audit/workplan_form.html'
    success_message = _("Workplan %(name)s was created successfully")
    
    def get_template_names(self):
        # Use modal template for HTMX requests
        if is_htmx_request(self.request):
            return ['audit/workplan_modal_form.html']
        return [self.template_name]
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add any additional context needed for CKEditor5 rendering
        return context
    
    def form_valid(self, form):
        try:
            # Set auto fields
            form.instance.organization = self.request.organization
            form.instance.created_by = self.request.user
            
            # Handle empty CKEditor fields - convert None to empty string
            if form.cleaned_data.get('objectives') is None:
                form.instance.objectives = ''
            if form.cleaned_data.get('description') is None:
                form.instance.description = ''
            
            # Ensure CKEditor content is properly sanitized
            if form.cleaned_data.get('objectives') == '<p>&nbsp;</p>':
                form.instance.objectives = ''
            if form.cleaned_data.get('description') == '<p>&nbsp;</p>':
                form.instance.description = ''
                
            return super().form_valid(form)
        except Exception as e:
            # Log the error but present a user-friendly message
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error creating workplan: {str(e)}", exc_info=True)
            form.add_error(None, _("An error occurred while creating the workplan. Please try again."))
            return self.form_invalid(form)
    
    def get_success_url(self):
        return reverse_lazy('audit:workplan-detail', kwargs={'pk': self.object.pk})

class WorkplanUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = AuditWorkplan
    form_class = AuditWorkplanForm
    template_name = 'audit/workplan_form.html'
    success_message = _("Workplan %(name)s was updated successfully")
    
    def get_template_names(self):
        # Use modal template for HTMX requests
        if is_htmx_request(self.request):
            return ['audit/workplan_modal_form.html']
        return [self.template_name]
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        return kwargs
    
    def form_valid(self, form):
        try:
            # Handle empty CKEditor fields - convert None to empty string
            if form.cleaned_data.get('objectives') is None:
                form.instance.objectives = ''
            if form.cleaned_data.get('description') is None:
                form.instance.description = ''
            
            # Ensure CKEditor content is properly sanitized
            if form.cleaned_data.get('objectives') == '<p>&nbsp;</p>':
                form.instance.objectives = ''
            if form.cleaned_data.get('description') == '<p>&nbsp;</p>':
                form.instance.description = ''
                
            return super().form_valid(form)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error updating workplan: {str(e)}", exc_info=True)
            form.add_error(None, _("An error occurred while updating the workplan. Please try again."))
            return self.form_invalid(form)
    
    def get_success_url(self):
        return reverse_lazy('audit:workplan-detail', kwargs={'pk': self.object.pk})

# ─── ENGAGEMENT VIEWS ────────────────────────────────────────────────────────
class EngagementListView(AuditPermissionMixin, ListView):
    model = Engagement
    template_name = 'audit/engagement_list.html'
    context_object_name = 'engagements'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        organization = self.request.organization
        queryset = queryset.filter(organization=organization).select_related(
            'annual_workplan', 'assigned_to', 'assigned_by'
        )
        form = EngagementFilterForm(self.request.GET)
        if form.is_valid():
            q = form.cleaned_data.get('q')
            if q:
                queryset = queryset.filter(Q(title__icontains=q) | Q(code__icontains=q))
            status = form.cleaned_data.get('status')
            if status:
                queryset = queryset.filter(project_status=status)
            owner = form.cleaned_data.get('owner')
            if owner:
                queryset = queryset.filter(assigned_to__email__icontains=owner)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = EngagementFilterForm(self.request.GET)
        return context

class EngagementDetailView(AuditPermissionMixin, DetailView):
    model = Engagement
    template_name = 'audit/engagement_detail.html'
    context_object_name = 'engagement'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        engagement = self.object
        # Aggregate all risks from all objectives under this engagement
        all_risks = Risk.objects.filter(objective__engagement=engagement)
        context.update({
            'issues': engagement.all_issues.select_related(
                'issue_owner', 'procedure_result'
            ),
            'approvals': engagement.approvals.all().select_related(
                'requester', 'approver'
            ),
            'can_submit': can_proceed(engagement.submit_for_approval),
            'can_approve': can_proceed(engagement.approve),
            'can_reject': can_proceed(engagement.reject),
            'content_type_id': ContentType.objects.get_for_model(type(engagement)).pk,
            'all_risks': all_risks,
        })
        return context

class EngagementCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = Engagement
    form_class = EngagementForm
    template_name = 'audit/engagement_form.html'
    success_message = _("Engagement %(title)s was created successfully")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        return kwargs
    
    def get_initial(self):
        initial = super().get_initial()
        if self.request.GET.get('workplan'):
            try:
                workplan_id = int(self.request.GET.get('workplan'))
                workplan = get_object_or_404(AuditWorkplan, pk=workplan_id, organization=self.request.organization)
                initial['annual_workplan'] = workplan
            except (ValueError, TypeError):
                pass
        return initial
    
    def get_template_names(self):
        """Return different templates based on request type."""
        if self.request.headers.get('HX-Request'):
            return ['audit/engagement_modal_form.html']
        return [self.template_name]
        
    def get(self, request, *args, **kwargs):
        # Check if this is an HTMX request - if so, we use the modal template via get_template_names()
        return super().get(request, *args, **kwargs)
    
    def form_valid(self, form):
        form.instance.organization = self.request.organization
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        # Handle HTMX or AJAX requests
        is_htmx = self.request.headers.get('HX-Request') == 'true'
        is_ajax = self.request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if is_htmx or is_ajax:
            from django.http import JsonResponse
            from django.urls import reverse
            detail_url = reverse('audit:engagement-detail', kwargs={'pk': self.object.pk})
            return JsonResponse({
                'success': True,
                'form_is_valid': True,  # Triggers successful form handling in modal-handler.js
                'pk': self.object.pk,
                'redirect': detail_url,  # Redirect to the engagement detail view
                'html_redirect': detail_url,  # For compatibility with existing JS
                'message': self.success_message % form.cleaned_data,
            })
        
        return response
    
    def form_invalid(self, form):
        """Handle form validation errors."""
        is_htmx = self.request.headers.get('HX-Request') == 'true'
        is_ajax = self.request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if is_htmx or is_ajax:
            from django.http import JsonResponse
            from django.template.loader import render_to_string
            
            # Render the form with errors
            html = render_to_string(
                self.get_template_names()[0],
                self.get_context_data(form=form),
                request=self.request
            )
            
            return JsonResponse({
                'success': False,
                'html': html,
            }, status=400)
        
        return super().form_invalid(form)
    
    def get_success_url(self):
        return reverse_lazy('audit:engagement-detail', kwargs={'pk': self.object.pk})

class EngagementUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = Engagement
    form_class = EngagementForm
    template_name = 'audit/engagement_form.html'
    success_message = _("Engagement %(title)s was updated successfully")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        return kwargs
    
    def get_template_names(self):
        """Return different templates based on request type."""
        if self.request.headers.get('HX-Request'):
            return ['audit/engagement_modal_form.html']
        return [self.template_name]
    
    def form_valid(self, form):
        """Handle successful form submission."""
        form.instance.organization = self.request.organization
        form.instance.last_modified_by = self.request.user
        response = super().form_valid(form)
        
        # Handle HTMX or AJAX requests
        is_htmx = self.request.headers.get('HX-Request') == 'true'
        is_ajax = self.request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if is_htmx or is_ajax:
            from django.http import JsonResponse
            from django.template.loader import render_to_string
            
            # Get the URL for the engagement list page
            from django.urls import reverse
            list_url = reverse('audit:engagement-list')
            
            # Provide JSON response for modal handler
            return JsonResponse({
                'success': True,
                'form_is_valid': True,  # Triggers successful form handling in modal-handler.js
                'pk': self.object.pk,
                'redirect': list_url,  # Always redirect to the engagement list
                'html_redirect': list_url,  # For compatibility with existing JS
                'message': self.success_message % form.cleaned_data,
            })
        
        return response
    
    def form_invalid(self, form):
        """Handle form validation errors."""
        is_htmx = self.request.headers.get('HX-Request') == 'true'
        is_ajax = self.request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if is_htmx or is_ajax:
            from django.http import JsonResponse
            from django.template.loader import render_to_string
            
            # Render the form with errors
            html = render_to_string(
                self.get_template_names()[0],
                self.get_context_data(form=form),
                request=self.request
            )
            
            return JsonResponse({
                'success': False,
                'html': html,
            }, status=400)
        
        return super().form_invalid(form)
    
    def get_success_url(self):
        return reverse_lazy('audit:engagement-detail', kwargs={'pk': self.object.pk})

# ─── APPROVAL WORKFLOW VIEWS ────────────────────────────────────────────────
@login_required
@permission_required('audit.can_submit_workplan', raise_exception=True)
def submit_workplan_for_approval(request, pk):
    """Submit a workplan for approval."""
    workplan = get_object_or_404(AuditWorkplan, pk=pk, organization=request.organization)
    
    # Check if the workplan is in draft state
    if workplan.approval_status != 'draft':
        messages.error(request, _('Workplan must be in draft status to submit for approval.'))
        return redirect('audit:workplan-detail', pk=workplan.pk)
    
    # Update approval status
    workplan.approval_status = 'submitted'
    workplan.save()
    
    # Create approval record
    Approval.objects.create(
        content_type=ContentType.objects.get_for_model(AuditWorkplan),
        object_id=workplan.pk,
        requester=request.user,
        organization=request.organization
    )
    
    # Send email notification
    try:
        from apps.audit.utils import send_workplan_approval_notification
    except ImportError:
        from .utils import send_workplan_approval_notification
    send_workplan_approval_notification(workplan, 'submitted', request)
    
    messages.success(request, _('Workplan has been submitted for approval.'), extra_tags='audit')
    return redirect('audit:workplan-detail', pk=workplan.pk)

@login_required
@permission_required('audit.can_approve_workplan', raise_exception=True)
def approve_workplan(request, pk):
    """Approve a workplan."""
    workplan = get_object_or_404(AuditWorkplan, pk=pk, organization=request.organization)
    
    # Check if the workplan is in submitted state
    if workplan.approval_status != 'submitted':
        messages.error(request, _('Workplan must be submitted for approval before it can be approved.'))
        return redirect('audit:workplan-detail', pk=workplan.pk)
    
    # Update approval status
    workplan.approval_status = 'approved'
    workplan.save()
    
    # Update approval record
    approval = Approval.objects.filter(
        content_type=ContentType.objects.get_for_model(AuditWorkplan),
        object_id=workplan.pk,
        status='pending'
    ).first()
    
    if approval:
        approval.status = 'approved'
        approval.approver = request.user
        approval.approved_at = timezone.now()
        approval.save()
    
    # Send email notification
    try:
        from apps.audit.utils import send_workplan_approval_notification
    except ImportError:
        from .utils import send_workplan_approval_notification
    send_workplan_approval_notification(workplan, 'approved', request)
    
    messages.success(request, _('Workplan has been approved.'), extra_tags='audit')
    return redirect('audit:workplan-detail', pk=workplan.pk)

@login_required
@permission_required('audit.can_approve_workplan', raise_exception=True)
def reject_workplan(request, pk):
    """Reject a workplan."""
    workplan = get_object_or_404(AuditWorkplan, pk=pk, organization=request.organization)
    
    # Check if the workplan is in submitted state
    if workplan.approval_status != 'submitted':
        messages.error(request, _('Workplan must be submitted for approval before it can be rejected.'))
        return redirect('audit:workplan-detail', pk=workplan.pk)
    
    # Update approval status
    workplan.approval_status = 'rejected'
    workplan.save()
    
    # Update approval record
    approval = Approval.objects.filter(
        content_type=ContentType.objects.get_for_model(AuditWorkplan),
        object_id=workplan.pk,
        status='pending'
    ).first()
    
    if approval:
        approval.status = 'rejected'
        approval.approver = request.user
        approval.approved_at = timezone.now()
        approval.save()
    
    # Send email notification
    try:
        from apps.audit.utils import send_workplan_approval_notification
    except ImportError:
        from .utils import send_workplan_approval_notification
    send_workplan_approval_notification(workplan, 'rejected', request)
    
    messages.success(request, _('Workplan has been rejected.'), extra_tags='audit')
    return redirect('audit:workplan-detail', pk=workplan.pk)

@login_required
@permission_required('audit.can_submit_engagement', raise_exception=True)
def submit_engagement_for_approval(request, pk):
    """Submit an engagement for approval."""
    engagement = get_object_or_404(Engagement, pk=pk, organization=request.organization)
    
    # Check if the engagement is in draft state
    if engagement.approval_status != 'draft':
        messages.error(request, _('Engagement must be in draft status to submit for approval.'))
        return redirect('audit:engagement-detail', pk=engagement.pk)
    
    # Update approval status
    engagement.approval_status = 'submitted'
    engagement.save()
    
    # Create approval record
    Approval.objects.create(
        content_type=ContentType.objects.get_for_model(Engagement),
        object_id=engagement.pk,
        requester=request.user,
        organization=request.organization
    )
    
    # Send email notification
    try:
        from apps.audit.utils import send_engagement_approval_notification
    except ImportError:
        from .utils import send_engagement_approval_notification
    send_engagement_approval_notification(engagement, 'submitted', request)
    
    messages.success(request, _('Engagement has been submitted for approval.'), extra_tags='audit')
    return redirect('audit:engagement-detail', pk=engagement.pk)

@login_required
@permission_required('audit.can_approve_engagement', raise_exception=True)
def approve_engagement(request, pk):
    """Approve an engagement."""
    engagement = get_object_or_404(Engagement, pk=pk, organization=request.organization)
    
    # Check if the engagement is in submitted state
    if engagement.approval_status != 'submitted':
        messages.error(request, _('Engagement must be submitted for approval before it can be approved.'))
        return redirect('audit:engagement-detail', pk=engagement.pk)
    
    # Update approval status
    engagement.approval_status = 'approved'
    engagement.save()
    
    # Update approval record
    approval = Approval.objects.filter(
        content_type=ContentType.objects.get_for_model(Engagement),
        object_id=engagement.pk,
        status='pending'
    ).first()
    
    if approval:
        approval.status = 'approved'
        approval.approver = request.user
        approval.approved_at = timezone.now()
        approval.save()
    
    # Send email notification
    try:
        from apps.audit.utils import send_engagement_approval_notification
    except ImportError:
        from .utils import send_engagement_approval_notification
    send_engagement_approval_notification(engagement, 'approved', request)
    
    messages.success(request, _('Engagement has been approved.'), extra_tags='audit')
    return redirect('audit:engagement-detail', pk=engagement.pk)

@login_required
@permission_required('audit.can_approve_engagement', raise_exception=True)
def reject_engagement(request, pk):
    """Reject an engagement."""
    engagement = get_object_or_404(Engagement, pk=pk, organization=request.organization)
    
    # Check if the engagement is in submitted state
    if engagement.approval_status != 'submitted':
        messages.error(request, _('Engagement must be submitted for approval before it can be rejected.'))
        return redirect('audit:engagement-detail', pk=engagement.pk)
    
    # Update approval status
    engagement.approval_status = 'rejected'
    engagement.save()
    
    # Update approval record
    approval = Approval.objects.filter(
        content_type=ContentType.objects.get_for_model(Engagement),
        object_id=engagement.pk,
        status='pending'
    ).first()
    
    if approval:
        approval.status = 'rejected'
        approval.approver = request.user
        approval.approved_at = timezone.now()
        approval.save()
    
    # Send email notification
    try:
        from apps.audit.utils import send_engagement_approval_notification
    except ImportError:
        from .utils import send_engagement_approval_notification
    send_engagement_approval_notification(engagement, 'rejected', request)
    
    messages.success(request, _('Engagement has been rejected.'), extra_tags='audit')
    return redirect('audit:engagement-detail', pk=engagement.pk)

# ─── ISSUE VIEWS ─────────────────────────────────────────────────────────────
class IssueListView(AuditPermissionMixin, ListView):
    model = Issue
    template_name = 'audit/issue_list.html'
    context_object_name = 'issues'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        organization = self.request.organization
        queryset = queryset.filter(organization=organization).select_related(
            'issue_owner', 'procedure__risk__objective__engagement'
        )
        form = IssueFilterForm(self.request.GET)
        if form.is_valid():
            q = form.cleaned_data.get('q')
            if q:
                queryset = queryset.filter(Q(issue_title__icontains=q) | Q(code__icontains=q))
            status = form.cleaned_data.get('status')
            if status:
                queryset = queryset.filter(issue_status=status)
            severity = form.cleaned_data.get('severity')
            if severity:
                queryset = queryset.filter(risk_level=severity)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = IssueFilterForm(self.request.GET)
        return context

class IssueDetailView(AuditPermissionMixin, DetailView):
    model = Issue
    template_name = 'audit/issue_detail.html'
    context_object_name = 'issue'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.contrib.contenttypes.models import ContentType
        issue = self.object
        # Robustly get engagement via procedure > risk > objective > engagement
        engagement = None
        if issue.procedure and hasattr(issue.procedure, 'risk') and hasattr(issue.procedure.risk, 'objective') and issue.procedure.risk.objective and hasattr(issue.procedure.risk.objective, 'engagement'):
            engagement = issue.procedure.risk.objective.engagement
        context['engagement'] = engagement
        # For notes
        content_type = ContentType.objects.get_for_model(issue)
        context['content_type_id'] = content_type.id
        context['object_id'] = issue.pk
        context['issue'] = issue  # For partials
        # Add any other context needed for modals/partials here
        context['working_papers'] = IssueWorkingPaper.objects.filter(issue=issue)
        return context

class IssueCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = Issue
    form_class = IssueForm
    template_name = 'audit/issue_form.html'
    success_message = _("Issue %(issue_title)s was created successfully")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        # Pass procedure_id from GET param to the form for auto-selection
        procedure_id = self.request.GET.get('procedure')
        if procedure_id:
            kwargs['procedure_id'] = procedure_id
        return kwargs
    
    def form_valid(self, form):
        form.instance.organization = self.request.organization
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        return super().form_valid(form)
    
    def get_success_url(self):
        # Redirect to issue detail after creation
        return reverse('audit:issue-detail', kwargs={'pk': self.object.pk})

class IssueUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = Issue
    form_class = IssueForm
    template_name = 'audit/issue_form.html'
    success_message = _("Issue %(issue_title)s was updated successfully")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        # Defensive: always check for self.object and its attributes
        try:
            if getattr(self, 'object', None) and getattr(self.object, 'procedure_result', None):
                kwargs['engagement_pk'] = self.object.procedure_result.procedure.objective.engagement_id
                kwargs['procedure_result_pk'] = self.object.procedure_result_id
        except Exception:
            # Never break, always return a dict
            pass
        return kwargs

# ─── APPROVAL VIEWS ──────────────────────────────────────────────────────────
class ApprovalCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = Approval
    form_class = ApprovalForm
    template_name = 'audit/approval/modal_form.html'
    success_message = _("Approval request was created successfully")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        kwargs['requester'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        form.instance.organization = self.request.organization
        form.instance.requester = self.request.user
        form.instance.status = 'pending'
        response = super().form_valid(form)
        
        # Send email notification to approver
        try:
            from apps.audit.utils import send_approval_notification
        except ImportError:
            from .utils import send_approval_notification
        send_approval_notification(self.object, self.request)
        
        # Handle HTMX request
        if is_htmx_request(self.request):
            context = {
                'success': True,
                'message': self.success_message % form.cleaned_data,
            }
            return render(self.request, 'audit/approval/success.html', context)
        
        return response
    
    def get_success_url(self):
        return reverse_lazy('audit:approval-detail', kwargs={'pk': self.object.pk})

class ApprovalDetailView(AuditPermissionMixin, DetailView):
    model = Approval
    template_name = 'audit/approval/detail.html'
    context_object_name = 'approval'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        approval = self.object
        
        # Add related data efficiently
        context.update({
            'content_object': approval.content_object,
            'can_approve': self.request.user == approval.approver,
            'can_reject': self.request.user == approval.approver,
            'content_type': ContentType.objects.get_for_model(approval.content_object).model,
            'organization': approval.organization,
        })
        return context

class PendingApprovalListView(AuditPermissionMixin, ListView):
    model = Approval
    template_name = 'audit/approval/pending_list.html'
    context_object_name = 'approvals'
    paginate_by = 20
    
    def get_queryset(self):
        # Only show pending approvals assigned to the current user
        return Approval.objects.filter(
            organization=self.request.organization,
            approver=self.request.user,
            status='pending'
        ).select_related('requester', 'approver').order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Pending Approvals')
        context['pending_count'] = self.get_queryset().count()
        return context

class ApprovalHistoryListView(AuditPermissionMixin, ListView):
    model = Approval
    template_name = 'audit/approval/history_list.html'
    context_object_name = 'approvals'
    paginate_by = 20
    
    def get_queryset(self):
        # Show approvals that the current user has approved or rejected
        return Approval.objects.filter(
            organization=self.request.organization,
            approver=self.request.user,
            status__in=['approved', 'rejected']
        ).select_related('requester', 'approver').order_by('-updated_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Approval History')
        context['history_count'] = self.get_queryset().count()
        return context

class RequestedApprovalsListView(AuditPermissionMixin, ListView):
    model = Approval
    template_name = 'audit/approval/requested_list.html'
    context_object_name = 'approvals'
    paginate_by = 20
    
    def get_queryset(self):
        # Show approvals requested by the current user
        return Approval.objects.filter(
            organization=self.request.organization,
            requester=self.request.user
        ).select_related('requester', 'approver').order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Requested Approvals')
        context['requested_count'] = self.get_queryset().count()
        return context

class ApprovalStatusUpdateView(AuditPermissionMixin, View):
    """View to handle approval or rejection of an approval request"""
    
    @transaction.atomic
    def post(self, request, pk):
        approval = get_object_or_404(Approval, pk=pk, organization=request.organization)
        
        # Only the assigned approver can approve/reject
        if request.user != approval.approver:
            raise PermissionDenied(_('You are not authorized to approve or reject this request'))
        
        # Can't update if already approved/rejected
        if approval.status != 'pending':
            messages.error(request, _('This approval has already been processed'))
            return redirect('audit:approval-detail', pk=approval.pk)
        
        action = request.POST.get('action')
        comments = request.POST.get('comments', '')
        
        if action == 'approve':
            approval.status = 'approved'
            approval.comments = comments
            approval.save()
            messages.success(request, _('Approval request has been approved'))
        elif action == 'reject':
            approval.status = 'rejected'
            approval.comments = comments
            approval.save()
            messages.success(request, _('Approval request has been rejected'))
        else:
            messages.error(request, _('Invalid action'))
            return redirect('audit:approval-detail', pk=approval.pk)
        
        # Send email notification to requester
        try:
            from apps.audit.utils import send_approval_status_notification
        except ImportError:
            from .utils import send_approval_status_notification
        send_approval_status_notification(approval, request)
        
        # Return appropriate response based on request type
        if is_htmx_request(request):
            context = {
                'success': True,
                'message': _('Approval status updated successfully'),
            }
            return render(request, 'audit/approval/success.html', context)
        
        return redirect('audit:approval-pending')

@login_required
def approve_approval(request, pk):
    """Function-based view for approval (for backward compatibility)"""
    approval = get_object_or_404(Approval, pk=pk, organization=request.organization)
    
    # Only the assigned approver can approve
    if request.user != approval.approver:
        raise PermissionDenied(_('You are not authorized to approve this request'))
    
    # Can't approve if already approved/rejected
    if approval.status != 'pending':
        messages.error(request, _('This approval has already been processed'))
        return redirect('audit:approval-detail', pk=approval.pk)
    
    with transaction.atomic():
        approval.status = 'approved'
        approval.comments = request.POST.get('comments', '')
        approval.save()
        
        # Send email notification to requester
        try:
            from apps.audit.utils import send_approval_status_notification
        except ImportError:
            from .utils import send_approval_status_notification
        send_approval_status_notification(approval, request)
    
    messages.success(request, _('Approval request has been approved'))
    return redirect('audit:approval-pending')

@login_required
def reject_approval(request, pk):
    """Function-based view for rejection (for backward compatibility)"""
    approval = get_object_or_404(Approval, pk=pk, organization=request.organization)
    
    # Only the assigned approver can reject
    if request.user != approval.approver:
        raise PermissionDenied(_('You are not authorized to reject this request'))
    
    # Can't reject if already approved/rejected
    if approval.status != 'pending':
        messages.error(request, _('This approval has already been processed'))
        return redirect('audit:approval-detail', pk=approval.pk)
    
    with transaction.atomic():
        approval.status = 'rejected'
        approval.comments = request.POST.get('comments', '')
        approval.save()
        
        # Send email notification to requester
        try:
            from apps.audit.utils import send_approval_status_notification
        except ImportError:
            from .utils import send_approval_status_notification
        send_approval_status_notification(approval, request)
    
    messages.success(request, _('Approval request has been rejected'))
    return redirect('audit:approval-pending')

# ─── DASHBOARD VIEW ──────────────────────────────────────────────────────────
class AuditDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'audit/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        organization = self.request.organization
        from collections import Counter, defaultdict
        from django.utils import timezone
        from .models.objective import Objective
        from .models.procedure import Procedure
        from .models.recommendation import Recommendation
        from .models.followupaction import FollowUpAction
        from .models.issueretest import IssueRetest
        from .models.note import Note
        from .models.issue import Issue
        from django.contrib.auth import get_user_model
        User = get_user_model()
        import calendar
        # --- Period Filter Logic ---
        org_created = organization.created_at.date()
        today = timezone.now().date()
        years = list(range(org_created.year, today.year + 1))
        months = [(i, calendar.month_name[i]) for i in range(1, 13)]
        selected_years = self.request.GET.getlist('year') or [str(today.year)]
        selected_months = self.request.GET.getlist('month')
        filter_all = 'All' in selected_years
        from django.db.models import Q
        period_q = Q()
        if not filter_all:
            year_ints = [int(y) for y in selected_years if y.isdigit()]
            if selected_months:
                month_ints = [int(m) for m in selected_months if m.isdigit()]
                period_q &= Q(project_start_date__year__in=year_ints, project_start_date__month__in=month_ints)
            else:
                period_q &= Q(project_start_date__year__in=year_ints)
        engagement_qs = Engagement.objects.filter(organization=organization)
        if not filter_all:
            engagement_qs = engagement_qs.filter(period_q)
        context['engagement_count'] = engagement_qs.count()
        context['recent_engagements'] = engagement_qs.order_by('-project_start_date')[:8]
        context['engagement_status_dist'] = dict(Counter(engagement_qs.values_list('project_status', flat=True))) or {}
        durations = [(e.target_end_date - e.project_start_date).days for e in engagement_qs if e.target_end_date and e.project_start_date]
        context['avg_engagement_duration'] = sum(durations) / len(durations) if durations else 0
        # Engagement Owner Workload: count by assigned_to full name (or email, or 'Unassigned')
        def owner_label(user):
            if user:
                if hasattr(user, 'get_full_name') and user.get_full_name():
                    return f"{user.get_full_name()} ({user.email})"
                elif hasattr(user, 'email'):
                    return user.email
            return 'Unassigned'
        owner_workload = [owner_label(e.assigned_to) for e in engagement_qs]
        context['engagement_owner_workload'] = dict(Counter(owner_workload)) or {}
        import json
        context['engagement_owner_workload_debug'] = json.dumps(context['engagement_owner_workload'], indent=2)
        # Issues: filter by procedure__risk__objective__engagement__in=engagement_qs
        issue_qs = Issue.objects.filter(
            organization=organization,
            procedure__risk__objective__engagement__in=engagement_qs
        )
        context['issue_count'] = issue_qs.count()
        context['recent_issues'] = issue_qs.order_by('-date_identified')[:8]
        context['overdue_issues'] = issue_qs.filter(issue_status__in=['open', 'in_progress'], target_date__lt=today).count()
        context['issue_risk_dist'] = dict(Counter(issue_qs.values_list('risk_level', flat=True))) or {}
        # Workplans filtered by period (if engagement_qs is filtered)
        workplan_ids = engagement_qs.values_list('annual_workplan_id', flat=True).distinct()
        workplan_qs = AuditWorkplan.objects.filter(organization=organization)
        if not filter_all:
            workplan_qs = workplan_qs.filter(id__in=workplan_ids)
        context['workplan_count'] = workplan_qs.count()
        context['recent_workplans'] = workplan_qs.order_by('-creation_date')[:8]
        completed_workplans = workplan_qs.filter(approval_status='approved').count()
        context['workplan_completion_rate'] = {'Completed': completed_workplans, 'Total': workplan_qs.count()}
        # Approvals filtered by period
        approval_qs = Approval.objects.filter(organization=organization)
        if not filter_all:
            approval_qs = approval_qs.filter(content_type__model='engagement', object_id__in=engagement_qs.values_list('id', flat=True))
        context['approval_count'] = approval_qs.count()
        context['approval_status_dist'] = dict(Counter(approval_qs.values_list('status', flat=True))) or {}
        context['pending_approvals'] = approval_qs.filter(status='pending').order_by('-created_at')[:8]
        context['available_years'] = years
        context['available_months'] = months
        context['selected_years'] = selected_years
        context['selected_months'] = selected_months
        context['filter_all'] = filter_all
        return context

class AuditWorkplanViewSet(viewsets.ModelViewSet):
    """API endpoint for audit workplans."""
    serializer_class = AuditWorkplanSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgManagerOrReadOnly]

    def get_queryset(self):
        return AuditWorkplan.objects.filter(organization=self.request.organization)

    def perform_create(self, serializer):
        serializer.save(organization=self.request.organization)

class EngagementViewSet(OrganizationScopedQuerysetMixin, viewsets.ModelViewSet):
    """API endpoint for audit engagements."""
    serializer_class = EngagementSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgManagerOrReadOnly]

    def perform_create(self, serializer):
        serializer.save(organization=self.request.organization)

class IssueViewSet(OrganizationScopedQuerysetMixin, viewsets.ModelViewSet):
    """API endpoint for audit issues."""
    serializer_class = IssueSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgManagerOrReadOnly]

    def perform_create(self, serializer):
        serializer.save(organization=self.request.organization)

class ApprovalViewSet(OrganizationScopedQuerysetMixin, viewsets.ModelViewSet):
    """API endpoint for audit approvals."""
    serializer_class = ApprovalSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgAdmin]

    def perform_create(self, serializer):
        serializer.save(
            organization=self.request.organization,
            requester=self.request.user
        )

class WorkplanEngagementViewSet(viewsets.ModelViewSet):
    """API endpoint for workplan engagements."""
    serializer_class = EngagementSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgManagerOrReadOnly]

    def get_queryset(self):
        workplan_pk = self.kwargs.get('workplan_pk')
        return Engagement.objects.filter(
            organization=self.request.organization,
            audit_workplan_id=workplan_pk
        )

    def perform_create(self, serializer):
        workplan_pk = self.kwargs.get('workplan_pk')
        workplan = AuditWorkplan.objects.get(pk=workplan_pk)
        serializer.save(
            organization=self.request.organization,
            audit_workplan=workplan
        )

class WorkplanApprovalViewSet(viewsets.ModelViewSet):
    """API endpoint for workplan approvals."""
    serializer_class = ApprovalSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgAdmin]

    def get_queryset(self):
        """Return approvals for the specified workplan."""
        workplan_id = self.kwargs.get('workplan_pk')
        return Approval.objects.filter(
            content_type__model='auditworkplan',
            object_id=workplan_id
        )

    def perform_create(self, serializer):
        """Create a new approval for the workplan."""
        workplan_id = self.kwargs.get('workplan_pk')
        workplan = get_object_or_404(AuditWorkplan, pk=workplan_id)
        serializer.save(
            content_type=ContentType.objects.get_for_model(AuditWorkplan),
            object_id=workplan_id,
            organization=workplan.organization
        )

class EngagementIssueViewSet(viewsets.ModelViewSet):
    """API endpoint for engagement issues."""
    serializer_class = IssueSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrgManagerOrReadOnly]
    
    def get_queryset(self):
        engagement_pk = self.kwargs.get('engagement_pk')
        return Issue.objects.filter(
            engagement_id=engagement_pk
        ).select_related('issue_owner', 'engagement')
    
    def perform_create(self, serializer):
        engagement_pk = self.kwargs.get('engagement_pk')
        engagement = get_object_or_404(Engagement, pk=engagement_pk)
        serializer.save(
            engagement=engagement,
            organization=self.request.organization,
            created_by=self.request.user
        )

class EngagementApprovalViewSet(viewsets.ModelViewSet):
    """API endpoint for engagement approvals."""
    permission_classes = [permissions.IsAuthenticated, IsOrgAdmin]
    
    def get_serializer_class(self):
        from .serializers import ApprovalSerializer
        return ApprovalSerializer
    
    def get_queryset(self):
        engagement_pk = self.kwargs.get('engagement_pk')
        return Approval.objects.filter(
            content_type__model='engagement',
            object_id=engagement_pk
        ).select_related('requester', 'approver')
    
    def perform_create(self, serializer):
        engagement_pk = self.kwargs.get('engagement_pk')
        engagement = get_object_or_404(Engagement, pk=engagement_pk)
        serializer.save(
            content_object=engagement,
            organization=self.request.organization,
            requester=self.request.user
        )


class EngagementRiskViewSet(viewsets.ModelViewSet):
    """API endpoint for engagement risks."""
    permission_classes = [permissions.IsAuthenticated, IsOrgManagerOrReadOnly]
    
    def get_queryset(self):
        engagement_pk = self.kwargs.get('engagement_pk')
        return Risk.objects.filter(
            engagement_id=engagement_pk
        ).select_related('created_by')
    
    def get_serializer_class(self):
        from .serializers import RiskSerializer, RiskDetailSerializer
        if self.action == 'retrieve':
            return RiskDetailSerializer
        return RiskSerializer
    
    def perform_create(self, serializer):
        engagement_pk = self.kwargs.get('engagement_pk')
        engagement = get_object_or_404(Engagement, pk=engagement_pk)
        serializer.save(
            engagement=engagement,
            organization=self.request.organization,
            created_by=self.request.user
        )


class ObjectiveRiskViewSet(viewsets.ModelViewSet):
    """API endpoint for objective risks."""
    permission_classes = [permissions.IsAuthenticated, IsOrgManagerOrReadOnly]
    
    def get_queryset(self):
        from .models.objective import Objective
        objective_pk = self.kwargs.get('objective_pk')
        return Risk.objects.filter(
            objective__id=objective_pk
        ).select_related('created_by')
    
    def get_serializer_class(self):
        from .serializers import RiskSerializer, RiskDetailSerializer
        if self.action == 'retrieve':
            return RiskDetailSerializer
        return RiskSerializer
    
    def perform_create(self, serializer):
        from .models.objective import Objective
        objective_pk = self.kwargs.get('objective_pk')
        objective = get_object_or_404(Objective, pk=objective_pk)
        serializer.save(
            objective=objective,
            organization=self.request.organization,
            created_by=self.request.user
        )

class IssueApprovalViewSet(viewsets.ModelViewSet):
    """API endpoint for issue approvals."""
    permission_classes = [permissions.IsAuthenticated, IsOrgAdmin]
    
    def get_serializer_class(self):
        from .serializers import ApprovalSerializer
        return ApprovalSerializer
    
    def get_queryset(self):
        issue_pk = self.kwargs.get('issue_pk')
        return Approval.objects.filter(
            content_type__model='issue',
            object_id=issue_pk
        ).select_related('requester', 'approver')
    
    def perform_create(self, serializer):
        issue_pk = self.kwargs.get('issue_pk')
        issue = get_object_or_404(Issue, pk=issue_pk)
        serializer.save(
            content_object=issue,
            organization=self.request.organization,
            requester=self.request.user
        )

@login_required
def submit_workplan(request, pk):
    workplan = get_object_or_404(AuditWorkplan, pk=pk)
    if can_proceed(workplan.submit_for_approval):
        workplan.submit_for_approval()
        workplan.save()
        messages.success(request, _("Workplan submitted for approval successfully."))
    else:
        messages.error(request, _("Cannot submit workplan for approval in its current state."))
    return redirect('audit:workplan-detail', pk=pk)

@login_required
def approve_workplan(request, pk):
    workplan = get_object_or_404(AuditWorkplan, pk=pk)
    if can_proceed(workplan.approve):
        workplan.approve()
        workplan.save()
        messages.success(request, _("Workplan approved successfully."))
    else:
        messages.error(request, _("Cannot approve workplan in its current state."))
    return redirect('audit:workplan-detail', pk=pk)

@login_required
def reject_workplan(request, pk):
    workplan = get_object_or_404(AuditWorkplan, pk=pk)
    if can_proceed(workplan.reject):
        workplan.reject()
        workplan.save()
        messages.success(request, _("Workplan rejected."))
    else:
        messages.error(request, _("Cannot reject workplan in its current state."))
    return redirect('audit:workplan-detail', pk=pk)

@login_required
def submit_engagement(request, pk):
    engagement = get_object_or_404(Engagement, pk=pk)
    if can_proceed(engagement.submit_for_approval):
        engagement.submit_for_approval()
        engagement.save()
        messages.success(request, _("Engagement submitted for approval successfully."))
    else:
        messages.error(request, _("Cannot submit engagement for approval in its current state."))
    return redirect('audit:engagement-detail', pk=pk)

@login_required
def approve_engagement(request, pk):
    engagement = get_object_or_404(Engagement, pk=pk)
    if can_proceed(engagement.approve):
        engagement.approve()
        engagement.save()
        messages.success(request, _("Engagement approved successfully."))
    else:
        messages.error(request, _("Cannot approve engagement in its current state."))
    return redirect('audit:engagement-detail', pk=pk)

@login_required
def reject_engagement(request, pk):
    engagement = get_object_or_404(Engagement, pk=pk)
    if can_proceed(engagement.reject):
        engagement.reject()
        engagement.save()
        messages.success(request, _("Engagement rejected."))
    else:
        messages.error(request, _("Cannot reject engagement in its current state."))
    return redirect('audit:engagement-detail', pk=pk)

@login_required
def close_issue(request, pk):
    issue = get_object_or_404(Issue, pk=pk)
    issue.issue_status = 'closed'
    issue.save()
    messages.success(request, _("Issue closed successfully."))
    return redirect('audit:issue-detail', pk=pk)

@login_required
def reopen_issue(request, pk):
    issue = get_object_or_404(Issue, pk=pk)
    issue.issue_status = 'open'
    issue.save()
    messages.success(request, _("Issue reopened successfully."))
    return redirect('audit:issue-detail', pk=pk)

@login_required
def approve_approval(request, pk):
    approval = get_object_or_404(Approval, pk=pk)
    approval.status = 'approved'
    approval.save()
    messages.success(request, _("Approval request approved successfully."))
    return redirect('audit:approval-detail', pk=pk)

@login_required
def reject_approval(request, pk):
    approval = get_object_or_404(Approval, pk=pk)
    approval.status = 'rejected'
    approval.save()
    messages.success(request, _("Approval request rejected."))
    return redirect('audit:approval-detail', pk=pk)

@login_required
def bulk_approve(request):
    if request.method == 'POST':
        ids = request.POST.getlist('selected_items')
        Approval.objects.filter(id__in=ids).update(status='approved')
        messages.success(request, _("Selected items approved successfully."))
    return redirect('audit:approval-list')

@login_required
def bulk_reject(request):
    if request.method == 'POST':
        ids = request.POST.getlist('selected_items')
        Approval.objects.filter(id__in=ids).update(status='rejected')
        messages.success(request, _("Selected items rejected successfully."))
    return redirect('audit:approval-list')

@login_required
def export_data(request):
    # Implement export logic here
    pass

@login_required
def workplan_report(request):
    # Implement workplan report generation here
    pass

@login_required
def engagement_report(request):
    # Implement engagement report generation here
    pass

@login_required
def issue_report(request):
    # Implement issue report generation here
    pass

@login_required
def approval_report(request):
    # Implement approval report generation here
    pass

@login_required
def search(request):
    # Implement search functionality here
    pass

@login_required
def autocomplete(request):
    # Implement autocomplete functionality here
    pass

@login_required
def validate(request):
    # Implement validation functionality here
    pass

def export_to_xlsx(headers, rows, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([smart_str(cell) for cell in row])
    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def export_workplans(request):
    fiscal_year = request.GET.get('fiscal_year')
    name = request.GET.get('name')
    export_format = request.GET.get('format', 'csv')
    qs = AuditWorkplan.objects.filter(organization=request.organization)
    if fiscal_year:
        qs = qs.filter(fiscal_year=fiscal_year)
    if name:
        qs = qs.filter(name__icontains=name)
    headers = ['ID', 'Code', 'Name', 'Fiscal Year', 'Approval Status', 'Creation Date']
    rows = [[wp.id, wp.code, wp.name, wp.fiscal_year, wp.approval_status, wp.creation_date] for wp in qs]
    if export_format == 'xlsx':
        return export_to_xlsx(headers, rows, 'workplans.xlsx')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="workplans.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response

@login_required
def export_engagements(request):
    qs = Engagement.objects.filter(organization=request.organization)
    # Apply the same filters as the list view
    q = request.GET.get('q')
    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(code__icontains=q))
    status = request.GET.get('status')
    if status:
        qs = qs.filter(project_status=status)
    owner = request.GET.get('owner')
    if owner:
        qs = qs.filter(assigned_to__email__icontains=owner)
    # Add more filters as needed

    export_format = request.GET.get('format', 'csv')
    headers = ['ID', 'Code', 'Title', 'Status', 'Assigned To', 'Start Date', 'End Date']
    rows = [[e.id, e.code, e.title, e.project_status, getattr(e.assigned_to, 'email', ''), e.project_start_date, e.target_end_date] for e in qs]
    if export_format == 'xlsx':
        return export_to_xlsx(headers, rows, 'engagements.xlsx')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="engagements.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response

@login_required
def export_issues(request):
    name = request.GET.get('name')
    severity = request.GET.get('severity')
    export_format = request.GET.get('format', 'csv')
    qs = Issue.objects.filter(organization=request.organization)
    if name:
        qs = qs.filter(issue_title__icontains=name)
    if severity:
        qs = qs.filter(risk_level=severity)
    headers = ['ID', 'Code', 'Title', 'Risk Level', 'Status', 'Owner', 'Date Identified']
    rows = [[i.id, i.code, i.issue_title, i.risk_level, i.issue_status, getattr(i.issue_owner, 'email', ''), i.date_identified] for i in qs]
    if export_format == 'xlsx':
        return export_to_xlsx(headers, rows, 'issues.xlsx')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="issues.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response

class ObjectiveListView(AuditPermissionMixin, ListView):
    model = Objective
    template_name = 'audit/objective_list.html'
    context_object_name = 'objectives'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        organization = self.request.organization
        return queryset.filter(engagement__organization=organization)

class ObjectiveDetailView(AuditPermissionMixin, DetailView):
    model = Objective
    template_name = 'audit/objective_detail.html'
    context_object_name = 'objective'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Use the correct related_name for risks
        context['risks'] = self.object.audit_risks.all()
        return context

class ObjectiveCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = Objective
    form_class = ObjectiveForm
    template_name = 'audit/objective_form.html'
    success_message = _('Objective %(title)s was created successfully')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        return kwargs

    def form_valid(self, form):
        engagement_pk = self.kwargs.get('engagement_pk')
        form.instance.engagement_id = engagement_pk
        form.instance.organization = self.request.organization
        response = super().form_valid(form)
        
        # Handle HTMX or AJAX requests
        is_htmx = self.request.headers.get('HX-Request') == 'true'
        is_ajax = self.request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if is_htmx or is_ajax:
            from django.http import JsonResponse
            from django.urls import reverse
            
            # Get the URL for the engagement detail page
            if engagement_pk:
                engagement_url = reverse('audit:engagement-detail', kwargs={'pk': engagement_pk})
            else:
                engagement_url = reverse('audit:objective-list')
            
            # Provide JSON response for modal handler
            return JsonResponse({
                'success': True,
                'form_is_valid': True,  # Triggers successful form handling in modal-handler.js
                'pk': self.object.pk,
                'redirect': engagement_url,
                'html_redirect': engagement_url,  # For compatibility with existing JS
                'message': self.success_message % form.cleaned_data,
            })
        
        return response

    def get_success_url(self):
        # Always redirect to the engagement detail view after creating an objective
        return reverse_lazy('audit:engagement-detail', kwargs={'pk': self.object.engagement.pk})

class ObjectiveUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = Objective
    form_class = ObjectiveForm
    template_name = 'audit/objective_form.html'
    success_message = _('Objective %(title)s was updated successfully')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        return kwargs

    def get_success_url(self):
        return reverse_lazy('audit:engagement-detail', kwargs={'pk': self.object.engagement.pk})

class ObjectiveModalCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = Objective
    form_class = ObjectiveForm
    template_name = 'audit/objective_modal_form.html'
    success_message = _('Objective %(title)s was created successfully')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['engagement_pk'] = self.kwargs.get('engagement_pk')
        return context

    def form_valid(self, form):
        engagement_pk = self.kwargs.get('engagement_pk')
        form.instance.engagement_id = engagement_pk
        form.instance.organization = self.request.organization
        response = super().form_valid(form)
        if self.request.htmx:
            return JsonResponse({
                'success': True,
                'redirect': reverse_lazy('audit:engagement-detail', kwargs={'pk': self.object.engagement.pk}),
            })
        return response

    def get_success_url(self):
        # Always redirect to the engagement detail view after creating an objective
        return reverse_lazy('audit:engagement-detail', kwargs={'pk': self.object.engagement.pk})

# PROCEDURE VIEWS
class ProcedureListView(AuditPermissionMixin, ListView):
    model = Procedure
    template_name = 'audit/procedure_list.html'
    context_object_name = 'procedures'
    paginate_by = 20
    def get_queryset(self):
        queryset = super().get_queryset()
        organization = self.request.organization
        return queryset.filter(objective__engagement__organization=organization)

class ProcedureDetailView(AuditPermissionMixin, DetailView):
    model = Procedure
    template_name = 'audit/procedure_detail.html'
    context_object_name = 'procedure'

class ProcedureCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = Procedure
    form_class = ProcedureForm
    template_name = 'audit/procedure_form.html'
    success_message = _('Procedure %(title)s was created successfully')
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        return kwargs
    def form_valid(self, form):
        objective_pk = self.kwargs.get('objective_pk')
        form.instance.objective_id = objective_pk
        form.instance.organization = self.request.organization
        return super().form_valid(form)
    def get_success_url(self):
        return reverse_lazy('audit:objective-detail', kwargs={'pk': self.object.objective.pk})

class ProcedureUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = Procedure
    form_class = ProcedureForm
    template_name = 'audit/procedure_form.html'
    success_message = _('Procedure %(title)s was updated successfully')
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        return kwargs
    def get_success_url(self):
        return reverse_lazy('audit:objective-detail', kwargs={'pk': self.object.objective.pk})

class ProcedureModalCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = Procedure
    form_class = ProcedureForm
    template_name = 'audit/procedure_modal_form.html'
    success_message = _('Procedure %(title)s was created successfully')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        kwargs['risk_id'] = self.kwargs.get('risk_id')
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['risk_id'] = self.kwargs.get('risk_id')
        return context
    
    def form_valid(self, form):
        risk_id = self.kwargs.get('risk_id')
        form.instance.risk_id = risk_id
        form.instance.organization = self.request.organization
        response = super().form_valid(form)
        # Handle HTMX requests
        if self.request.headers.get('HX-Request') == 'true':
            procedures = self.model.objects.filter(risk_id=risk_id).order_by('order', 'id')
            html_list = render_to_string('audit/_procedure_list_partial.html', {
                'procedures': procedures,
                'risk': form.instance.risk,
            }, request=self.request)
            return JsonResponse({'success': True, 'pk': self.object.pk, 'title': self.object.title, 'html_list': html_list})
        return response

    def get_success_url(self):
        return reverse_lazy('audit:risk-detail', kwargs={'pk': self.object.risk.pk})

class ProcedureModalUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = Procedure
    form_class = ProcedureForm
    template_name = 'audit/procedure_modal_form.html'
    success_message = _('Procedure %(title)s was updated successfully')
    
    def get_queryset(self):
        """Limit queryset to objects in the current organization."""
        return super().get_queryset().filter(organization=self.request.organization)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        if hasattr(self, 'object') and self.object.risk_id:
            kwargs['risk_id'] = self.object.risk_id
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if hasattr(self, 'object') and self.object.risk:
            context['risk'] = self.object.risk
        return context

    def form_valid(self, form):
        form.instance.organization = self.request.organization
        response = super().form_valid(form)
        
        # Handle HTMX/JSON response
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest" or self.request.headers.get("HX-Request") == "true":
            try:
                # Return a success response that will close the modal
                return JsonResponse({
                    'success': True,
                    'message': str(self.success_message),
                    'redirect_url': self.get_success_url()
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': _("Error updating procedure: {}".format(str(e)))
                }, status=400)
        
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Create a basic filter form structure
        from django import forms
        from crispy_forms.helper import FormHelper
        from crispy_forms.layout import Submit
        
        class EmptyFilterForm(forms.Form):
            # Empty form - just to prevent the template error
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self.helper = FormHelper()
                self.helper.form_method = 'get'
                self.helper.add_input(Submit('submit', 'Filter', css_class='btn-secondary'))
        
        context['filter_form'] = EmptyFilterForm()
            
        return context



class ProcedureModalDeleteView(AuditPermissionMixin, DeleteView):
    model = Procedure
    template_name = 'audit/procedure_confirm_delete.html'
    success_message = _('Procedure was deleted successfully')

    def get_queryset(self):
        """Limit queryset to objects in the current organization."""
        return super().get_queryset().filter(organization=self.request.organization)
    
    def get_success_url(self):
        """Return to the risk detail page after deletion."""
        if hasattr(self, 'object') and self.object.risk_id:
            return reverse_lazy('audit:risk-detail', kwargs={'pk': self.object.risk_id})
        return reverse_lazy('audit:procedure-list')

    def delete(self, request, *args, **kwargs):
        """Handle the deletion and return appropriate response."""
        self.object = self.get_object()
        success_url = self.get_success_url()
        
        # Store risk_id before deletion for HTMX response
        risk_id = self.object.risk_id if self.object.risk else None
        
        # Perform deletion
        self.object.delete()
        
        # Handle HTMX/JSON response
        if request.headers.get("x-requested-with") == "XMLHttpRequest" or request.headers.get("HX-Request") == "true":
            try:
                if risk_id:
                    # Get updated procedure list for the risk
                    from .models.risk import Risk
                    risk = Risk.objects.get(pk=risk_id, organization=request.organization)
                    procedures = Procedure.objects.filter(risk=risk, organization=request.organization)
                    
                    # Render the procedure list partial
                    html = render_to_string(
                        'audit/_procedure_list_partial.html',
                        {
                            'procedures': procedures,
                            'risk': risk,
                            'request': request
                        }
                    )
                    
                    return HttpResponse(html)
                
                return JsonResponse({
                    'success': True,
                    'message': str(self.success_message),
                    'redirect_url': success_url
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': _("Error deleting procedure: {}".format(str(e)))
                }, status=400)
        
        messages.success(request, self.success_message)
        return HttpResponseRedirect(success_url)

# FOLLOWUP ACTION VIEWS
class FollowUpActionListView(AuditPermissionMixin, ListView):
    model = FollowUpAction
    template_name = 'audit/followupaction_list.html'
    context_object_name = 'followup_actions'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        organization = self.request.organization
        
        # Filter by organization
        queryset = queryset.filter(organization=organization)
        
        # Filter by issue if specified
        issue_pk = self.request.GET.get('issue_pk')
        if issue_pk:
            queryset = queryset.filter(issue_id=issue_pk)
            
        # Filter by status if specified
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        # Filter by assigned_to if specified
        assigned_to = self.request.GET.get('assigned_to')
        if assigned_to:
            queryset = queryset.filter(assigned_to_id=assigned_to)
            
        # Order by due date (ascending, with nulls last)
        queryset = queryset.order_by(
            models.F('due_date').asc(nulls_last=True),
            '-created_at'
        )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        organization = self.request.organization
        
        # Add filter form data
        context['status_choices'] = dict(FollowUpAction.STATUS_CHOICES)
        
        # Add issue info if filtered by issue
        issue_pk = self.request.GET.get('issue_pk')
        if issue_pk:
            from .models.issue import Issue
            try:
                context['issue'] = Issue.objects.get(pk=issue_pk, organization=organization)
            except Issue.DoesNotExist:
                pass
                
        # Add assigned users for filter
        context['assigned_users'] = CustomUser.objects.filter(
            organization=organization,
            followup_actions_assigned__isnull=False
        ).distinct()
        
        return context

class FollowUpActionDetailView(AuditPermissionMixin, DetailView):
    model = FollowUpAction
    template_name = 'audit/followupaction_detail.html'
    context_object_name = 'followup_action'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(organization=self.request.organization)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_edit'] = self.request.user.has_perm('audit.change_followupaction')
        context['can_delete'] = self.request.user.has_perm('audit.delete_followupaction')
        return context

class FollowUpActionCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = FollowUpAction
    form_class = FollowUpActionForm
    template_name = 'audit/followupaction_form.html'
    success_message = _("Follow Up Action was created successfully")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        # Get issue_pk from URL or request data
        issue_pk = self.kwargs.get('issue_pk') or self.request.GET.get('issue_pk')
        if issue_pk:
            kwargs['issue_pk'] = issue_pk
        return kwargs
    
    def get_success_url(self):
        # Redirect to parent issue detail
        if self.object.issue:
            return reverse('audit:issue-detail', kwargs={'pk': self.object.issue.pk})
        return reverse('audit:followupaction-list')
    
    def get_initial(self):
        initial = super().get_initial()
        # Get issue_pk from URL or request data
        issue_pk = self.kwargs.get('issue_pk') or self.request.GET.get('issue_pk')
        if issue_pk:
            initial['issue'] = issue_pk
        return initial
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get issue_pk from URL or request data
        issue_pk = self.kwargs.get('issue_pk') or self.request.GET.get('issue_pk')
        if issue_pk:
            try:
                context['issue'] = Issue.objects.get(pk=issue_pk, organization=self.request.organization)
            except Issue.DoesNotExist:
                messages.error(self.request, _('Invalid issue selected'))
        return context
    
    def form_valid(self, form):
        # Defensive: always set organization before saving
        form.instance.organization = self.request.organization
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, _('Please correct the errors below.'))
        return super().form_invalid(form)

class FollowUpActionUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = FollowUpAction
    form_class = FollowUpActionForm
    template_name = 'audit/followupaction_form.html'
    success_message = _("Follow Up Action was updated successfully")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["organization"] = self.request.organization
        if hasattr(self, 'object') and self.object.issue_id:
            kwargs["issue_pk"] = self.object.issue_id
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if hasattr(self, 'object') and self.object.issue:
            context["issue"] = self.object.issue
            context["issue_pk"] = self.object.issue_id
        return context
    
    def form_valid(self, form):
        form.instance.organization = self.request.organization
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest" or self.request.headers.get("HX-Request") == "true":
            from .models.issue import Issue
            followups = FollowUpAction.objects.filter(issue_id=self.object.issue_id, organization=self.request.organization)
            html_list = render_to_string("audit/followupaction_list.html", {
                "followup_actions": followups,
                "issue": Issue.objects.get(pk=self.object.issue_id),
            }, request=self.request)
            return JsonResponse({"form_is_valid": True, "html_list": html_list})
        return response
    
    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest" or self.request.headers.get("HX-Request") == "true":
            html_form = render_to_string(self.template_name, {"form": form}, request=self.request)
            return JsonResponse({"form_is_valid": False, "html_form": html_form})
        return super().form_invalid(form)
    
    def get_success_url(self):
        if hasattr(self, 'object') and self.object.issue_id:
            return reverse_lazy('audit:issue-detail', kwargs={'pk': self.object.issue_id})
        return reverse_lazy('audit:followupaction-list')

class FollowUpActionModalCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = FollowUpAction
    form_class = FollowUpActionForm
    template_name = 'audit/followupaction_modal_form.html'
    success_message = _('Follow Up Action was created successfully')

    def get_issue_pk(self):
        """Get the issue_pk from URL kwargs or request parameters."""
        return self.kwargs.get("issue_pk") or self.request.GET.get("issue_pk")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["organization"] = self.request.organization
        issue_pk = self.get_issue_pk()
        if issue_pk:
            kwargs["issue_pk"] = issue_pk
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        issue_pk = self.get_issue_pk()
        if issue_pk:
            from .models.issue import Issue
            try:
                context["issue"] = Issue.objects.get(pk=issue_pk, organization=self.request.organization)
            except Issue.DoesNotExist:
                pass
        return context

    def get(self, request, *args, **kwargs):
        """Handle GET request to display the form."""
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        
        # Add HTMX headers to the response
        context = self.get_context_data(form=form)
        if request.headers.get('HX-Request'):
            return self.render_to_response(context)
        return super().get(request, *args, **kwargs)
        
    def form_valid(self, form):
        """Handle form validation and submission."""
        issue_pk = self.get_issue_pk()
        if not issue_pk:
            form.add_error(None, _("Issue is required for creating a follow-up action."))
            return self.form_invalid(form)
            
        form.instance.issue_id = issue_pk
        form.instance.organization = self.request.organization
        form.instance.created_by = self.request.user
        
        try:
            self.object = form.save()
            
            # Handle HTMX request
            if self.request.headers.get('HX-Request'):
                from django.template.loader import render_to_string
                from .models.issue import Issue
                
                try:
                    issue = Issue.objects.get(pk=issue_pk, organization=self.request.organization)
                    followups = FollowUpAction.objects.filter(issue=issue, organization=self.request.organization)
                    
                    # Render the follow-up list partial
                    html = render_to_string(
                        'audit/_followupaction_list_partial.html',
                        {
                            'followups': followups,
                            'issue': issue,
                            'request': self.request
                        }
                    )
                    
                    # Return success response with HTML
                    return HttpResponse(html)
                    
                except Issue.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'message': _("Error: Related issue not found.")
                    }, status=400)
            
            # For non-HTMX requests, redirect to success URL
            messages.success(self.request, self.get_success_message(form.cleaned_data))
            return redirect(self.get_success_url())
            
        except Exception as e:
            if self.request.headers.get('HX-Request'):
                return JsonResponse({
                    'success': False,
                    'message': str(e)
                }, status=400)
            raise

    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest" or self.request.headers.get("HX-Request") == "true":
            html_form = render_to_string(
                self.template_name,
                self.get_context_data(form=form),
                request=self.request
            )
            return JsonResponse({
                "form_is_valid": False,
                "html_form": html_form
            })
        return super().form_invalid(form)

    def get_success_url(self):
        issue_pk = self.get_issue_pk()
        if issue_pk:
            return reverse_lazy("audit:issue-detail", kwargs={"pk": issue_pk})
        return reverse_lazy("audit:followupaction-list")

class FollowUpActionModalUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = FollowUpAction
    form_class = FollowUpActionForm
    template_name = 'audit/followupaction_modal_form.html'
    success_message = _('Follow Up Action was updated successfully')

    def get_queryset(self):
        """Limit queryset to objects in the current organization."""
        return super().get_queryset().filter(organization=self.request.organization)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        if hasattr(self, 'object') and self.object.issue_id:
            kwargs['issue_pk'] = self.object.issue_id
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if hasattr(self, 'object') and self.object.issue:
            context['issue'] = self.object.issue
        return context

    def form_valid(self, form):
        form.instance.organization = self.request.organization
        response = super().form_valid(form)
        
        # Handle HTMX/JSON response
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest" or self.request.headers.get("HX-Request") == "true":
            try:
                # Return a success response that will close the modal
                return JsonResponse({
                    'success': True,
                    'message': str(self.success_message),
                    'redirect_url': self.get_success_url()
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': _("Error updating follow-up action: {}".format(str(e)))
                }, status=400)
        
        return response

    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest" or self.request.headers.get("HX-Request") == "true":
            html_form = render_to_string(
                self.template_name,
                self.get_context_data(form=form),
                request=self.request
            )
            return JsonResponse({
                'success': False,
                'form_errors': form.errors,
                'html_form': html_form
            }, status=400)
        return super().form_invalid(form)

    def get_success_url(self):
        if hasattr(self, 'object') and self.object.issue_id:
            return reverse_lazy('audit:issue-detail', kwargs={'pk': self.object.issue_id})
        return reverse_lazy('audit:followupaction-list')

class FollowUpActionDeleteView(AuditPermissionMixin, DeleteView):
    model = FollowUpAction
    template_name = 'audit/followupaction_confirm_delete.html'
    success_message = _('Follow-up action was deleted successfully')

    def get_queryset(self):
        """Limit queryset to objects in the current organization."""
        return super().get_queryset().filter(organization=self.request.organization)

    def get_success_url(self):
        """Return to the issue detail page after deletion."""
        if hasattr(self, 'object') and self.object.issue_id:
            return reverse_lazy('audit:issue-detail', kwargs={'pk': self.object.issue_id})
        return reverse_lazy('audit:followupaction-list')

    def delete(self, request, *args, **kwargs):
        """Handle the deletion and return appropriate response."""
        self.object = self.get_object()
        success_url = self.get_success_url()
        
        # Store issue_id before deletion for HTMX response
        issue_id = self.object.issue_id if self.object.issue else None
        
        # Perform deletion
        self.object.delete()
        
        # Handle HTMX/JSON response
        if request.headers.get("x-requested-with") == "XMLHttpRequest" or request.headers.get("HX-Request") == "true":
            try:
                if issue_id:
                    # Get updated follow-up list for the issue
                    from .models.issue import Issue
                    issue = Issue.objects.get(pk=issue_id, organization=request.organization)
                    followups = FollowUpAction.objects.filter(issue=issue, organization=request.organization)
                    
                    # Render the follow-up list partial
                    html = render_to_string(
                        'audit/_followupaction_list_partial.html',
                        {
                            'followups': followups,
                            'issue': issue,
                            'request': request
                        }
                    )
                    
                    return HttpResponse(html)
                
                return JsonResponse({
                    'success': True,
                    'message': str(self.success_message),
                    'redirect_url': success_url
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': _("Error deleting follow-up action: {}".format(str(e)))
                }, status=400)
        
        messages.success(request, self.success_message)
        return HttpResponseRedirect(success_url)

# ISSUE RETEST VIEWS
class IssueRetestListView(AuditPermissionMixin, ListView):
    model = IssueRetest
    template_name = 'audit/issueretest_list.html'
    context_object_name = 'issue_retests'
    paginate_by = 20
    def get_queryset(self):
        queryset = super().get_queryset()
        organization = self.request.organization
        return queryset.filter(issue__procedure_result__procedure__objective__engagement__organization=organization)

class IssueRetestDetailView(AuditPermissionMixin, DetailView):
    model = IssueRetest
    template_name = 'audit/issueretest_detail.html'
    context_object_name = 'issue_retest'

class IssueRetestCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = IssueRetest
    form_class = IssueRetestForm
    template_name = 'audit/issueretest_form.html'
    success_message = _("Issue Retest was created successfully")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        # Get issue_pk from URL or request data
        issue_pk = self.kwargs.get('issue_pk') or self.request.GET.get('issue_pk')
        if issue_pk:
            kwargs['issue_pk'] = issue_pk
        return kwargs
    
    def get_success_url(self):
        # Redirect to parent issue detail
        if self.object.issue:
            return reverse('audit:issue-detail', kwargs={'pk': self.object.issue.pk})
        return reverse('audit:issueretest-list')
    
    def get_initial(self):
        initial = super().get_initial()
        # Get issue_pk from URL or request data
        issue_pk = self.kwargs.get('issue_pk') or self.request.GET.get('issue_pk')
        if issue_pk:
            initial['issue'] = issue_pk
        return initial
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get issue_pk from URL or request data
        issue_pk = self.kwargs.get('issue_pk') or self.request.GET.get('issue_pk')
        if issue_pk:
            try:
                context['issue'] = Issue.objects.get(pk=issue_pk, organization=self.request.organization)
            except Issue.DoesNotExist:
                messages.error(self.request, _('Invalid issue selected'))
        return context
    
    def form_valid(self, form):
        # Defensive: always set organization before saving
        form.instance.organization = self.request.organization
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, _('Please correct the errors below.'))
        return super().form_invalid(form)

class IssueRetestUpdateView(IssueRetestCreateView, UpdateView):
    success_message = _("Issue Retest was updated successfully")

class IssueRetestModalCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = IssueRetest
    form_class = IssueRetestForm
    template_name = 'audit/issueretest_modal_form.html'
    success_message = _('Retest was created successfully')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        kwargs['issue_pk'] = self.kwargs.get('issue_pk')
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        issue_pk = self.kwargs.get('issue_pk')
        context['issue_pk'] = issue_pk
        
        # Add the issue object to context
        if issue_pk:
            try:
                from .models.issue import Issue
                context['issue'] = Issue.objects.get(pk=issue_pk, organization=self.request.organization)
            except Issue.DoesNotExist:
                context['issue'] = None
        else:
            context['issue'] = None
            
        return context

    def form_valid(self, form):
        issue_pk = self.kwargs.get('issue_pk')
        form.instance.issue_id = issue_pk
        form.instance.organization = self.request.organization
        response = super().form_valid(form)
        
        # For HTMX requests, return JSON with updated list
        if self.request.htmx or self.request.headers.get('HX-Request') == 'true':
            from django.template.loader import render_to_string
            from django.http import JsonResponse
            try:
                issue_obj = Issue.objects.get(pk=issue_pk)
            except Issue.DoesNotExist:
                issue_obj = None
            html = render_to_string('audit/_issueretest_list_partial.html', {
                'retests': IssueRetest.objects.filter(issue_id=issue_pk),
                'issue': issue_obj
            }, request=self.request)
            return JsonResponse({
                'form_is_valid': True,
                'html_list': html,
                'message': 'Retest was created successfully.'
            })
        return response

    def form_invalid(self, form):
        if self.request.htmx or self.request.headers.get('HX-Request') == 'true':
            from django.template.loader import render_to_string
            from django.http import JsonResponse
            html = render_to_string(self.template_name, self.get_context_data(form=form), request=self.request)
            return JsonResponse({
                'form_is_valid': False,
                'html_form': html
            }, status=400)
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('audit:issue-detail', kwargs={'pk': self.object.issue.pk})

class IssueRetestModalUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = IssueRetest
    form_class = IssueRetestForm
    template_name = 'audit/issueretest_modal_form.html'
    success_message = _('Retest was updated successfully')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        kwargs['issue_pk'] = self.object.issue.pk if self.object.issue_id else None
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.object.issue_id:
            context['issue'] = self.object.issue
        else:
            context['issue'] = None
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.htmx or self.request.headers.get('HX-Request') == 'true':
            from django.template.loader import render_to_string
            from django.http import JsonResponse
            html = render_to_string('audit/_issueretest_list_partial.html', {
                'retests': IssueRetest.objects.filter(issue_id=self.object.issue_id),
                'issue': self.object.issue
            }, request=self.request)
            return JsonResponse({
                'form_is_valid': True,
                'html_list': html,
                'message': 'Retest was updated successfully.'
            })
        return response

    def form_invalid(self, form):
        if self.request.htmx or self.request.headers.get('HX-Request') == 'true':
            from django.template.loader import render_to_string
            from django.http import JsonResponse
            html = render_to_string(self.template_name, self.get_context_data(form=form), request=self.request)
            return JsonResponse({
                'form_is_valid': False,
                'html_form': html
            }, status=400)
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('audit:issue-detail', kwargs={'pk': self.object.issue.pk})

# NOTE VIEWS (GENERIC, MODAL)
class NoteCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = Note
    form_class = NoteForm
    template_name = 'audit/note_modal_form.html'
    success_message = _('Note was created successfully')

    def get_issue_pk(self):
        issue_pk = self.kwargs.get('issue_pk') or self.request.GET.get('issue_pk') or self.request.POST.get('issue_pk')
        return issue_pk

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        # Only pass issue_pk if it exists
        issue_pk = self.get_issue_pk()
        if issue_pk:
            kwargs['issue_pk'] = issue_pk
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add content type and object id to the context for proper rendering
        context['content_type_id'] = self.kwargs.get('content_type_id')
        context['object_id'] = self.kwargs.get('object_id')
        
        # Add issue_pk only if it exists
        issue_pk = self.get_issue_pk()
        if issue_pk:
            context['issue_pk'] = issue_pk
        
        # Robustly resolve engagement_pk for cancel/back button
        engagement_pk = None
        object_id = self.kwargs.get('object_id')
        content_type_id = self.kwargs.get('content_type_id')
        if object_id and content_type_id:
            from django.contrib.contenttypes.models import ContentType
            ct = ContentType.objects.get_for_id(content_type_id)
            model_class = ct.model_class()
            try:
                obj = model_class.objects.get(pk=object_id)
                # Try to resolve engagement from the object
                if hasattr(obj, 'engagement_id') and obj.engagement_id:
                    engagement_pk = obj.engagement_id
                elif hasattr(obj, 'engagement') and obj.engagement:
                    engagement_pk = obj.engagement.pk
                elif hasattr(obj, 'objective') and obj.objective and hasattr(obj.objective, 'engagement_id'):
                    engagement_pk = obj.objective.engagement_id
                elif hasattr(obj, 'procedure') and obj.procedure and hasattr(obj.procedure, 'objective') and obj.procedure.objective and hasattr(obj.procedure.objective, 'engagement_id'):
                    engagement_pk = obj.procedure.objective.engagement_id
                # Add more logic as needed for other models
            except Exception:
                pass
        context['engagement_pk'] = engagement_pk
        return context

    def form_valid(self, form):
        form.instance.organization = self.request.organization
        form.instance.content_type_id = self.kwargs.get('content_type_id')
        form.instance.object_id = self.kwargs.get('object_id')
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest" or self.request.headers.get("HX-Request") == "true":
            notes = Note.objects.filter(content_type_id=self.kwargs.get('content_type_id'), object_id=self.kwargs.get('object_id'), organization=self.request.organization)
            html_list = render_to_string("audit/_note_list_partial.html", {
                "notes": notes,
            }, request=self.request)
            return JsonResponse({"form_is_valid": True, "html_list": html_list})
        return response

    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            html_form = render_to_string(self.template_name, {"form": form}, request=self.request)
            return JsonResponse({"form_is_valid": False, "html_form": html_form})
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('audit:note-list')

class NoteModalUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = Note
    form_class = NoteForm
    template_name = "audit/note_modal_form.html"
    success_message = _("Note was updated successfully")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['content_type_id'] = self.object.content_type_id
        context['object_id'] = self.object.object_id
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest" or self.request.headers.get("HX-Request") == "true":
            notes = Note.objects.filter(content_type_id=self.object.content_type_id, object_id=self.object.object_id, organization=self.request.organization)
            html_list = render_to_string("audit/_note_list_partial.html", {
                "notes": notes,
            }, request=self.request)
            return JsonResponse({"form_is_valid": True, "html_list": html_list})
        return response

    def form_invalid(self, form):
        if self.request.headers.get("x-requested-with") == "XMLHttpRequest":
            html_form = render_to_string(self.template_name, {"form": form}, request=self.request)
            return JsonResponse({"form_is_valid": False, "html_form": html_form})
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('audit:note-list')

class NotificationListView(generics.ListAPIView):
    serializer_class = NotificationSerializer
    permission_classes = []  # Allow unauthenticated access
    
    def get_queryset(self):
        # Return empty queryset for unauthenticated users
        if not self.request.user.is_authenticated:
            return Notification.objects.none()
        # Only return notifications for authenticated users
        return Notification.objects.filter(user=self.request.user, user__organization=getattr(self.request, 'organization', None))

class NotificationTemplateView(AuditPermissionMixin, ListView):
    """A template-based view for displaying notifications in a user-friendly UI"""
    model = Notification
    template_name = 'audit/notification_list.html'
    context_object_name = 'notifications'
    paginate_by = 20
    
    def get_queryset(self):
        # Get notifications for the current user and mark them as read
        queryset = Notification.objects.filter(
            user=self.request.user, 
            user__organization=self.request.organization
        ).order_by('-created_at')
        
        # Mark all as read (optional, remove if you want to keep unread status)
        # queryset.update(is_read=True)
        
        return queryset
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unread_count'] = Notification.objects.filter(
            user=self.request.user,
            user__organization=self.request.organization,
            is_read=False
        ).count()
        return context

class NoteListView(AuditPermissionMixin, ListView):
    model = Note
    template_name = 'audit/note_list.html'
    context_object_name = 'notes'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        organization = self.request.organization
        queryset = queryset.filter(organization=organization)
        # Optionally filter by content_type/object_id if provided
        content_type_id = self.request.GET.get('content_type_id')
        object_id = self.request.GET.get('object_id')
        if content_type_id and object_id:
            queryset = queryset.filter(content_type_id=content_type_id, object_id=object_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['content_type_id'] = self.request.GET.get('content_type_id')
        context['object_id'] = self.request.GET.get('object_id')
        return context

class RecommendationListView(AuditPermissionMixin, ListView):
    model = Recommendation
    template_name = "audit/recommendation_list.html"
    context_object_name = "recommendations"

    def get_queryset(self):
        issue_pk = self.kwargs["issue_pk"]
        return Recommendation.objects.filter(issue_id=issue_pk, organization=self.request.organization)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models.issue import Issue
        context["issue"] = Issue.objects.get(pk=self.kwargs["issue_pk"])
        return context

class RecommendationCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = Recommendation
    form_class = RecommendationForm
    template_name = "audit/recommendation_form.html"
    success_message = _("Recommendation was created successfully")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        issue_pk = safe_get_issue_pk(self)
        if issue_pk:
            kwargs['issue_pk'] = issue_pk
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get issue_pk from URL or request data
        issue_pk = self.kwargs.get('issue_pk') or self.request.GET.get('issue_pk')
        if issue_pk:
            try:
                context['issue'] = Issue.objects.get(pk=issue_pk, organization=self.request.organization)
            except Issue.DoesNotExist:
                messages.error(self.request, _('Invalid issue selected'))
        return context
    
    def form_valid(self, form):
        # Defensive: always set organization before saving
        form.instance.organization = self.request.organization
        return super().form_valid(form)
    
    def get_success_url(self):
        if self.object.issue:
            return reverse('audit:issue-detail', kwargs={'pk': self.object.issue.pk})
        return reverse('audit:recommendation-list')

class RecommendationUpdateView(RecommendationCreateView, UpdateView):
    success_message = _("Recommendation was updated successfully")
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        # Do NOT set 'issue_pk' for edit
        return kwargs
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['issue'] = self.object.issue  # Ensure 'issue' is always available for the modal
        return context

class RecommendationDetailView(AuditPermissionMixin, DetailView):
    model = Recommendation
    template_name = "audit/recommendation_detail.html"
    context_object_name = "recommendation"

class RecommendationModalUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = Recommendation
    form_class = RecommendationForm
    template_name = 'audit/recommendation_modal_form.html'
    success_message = _('Recommendation was updated successfully')
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['organization'] = self.request.organization
        # Do NOT set 'issue_pk' for edit
        return kwargs
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['issue'] = self.object.issue  # Ensure 'issue' is always available for the modal
        return context

@login_required
def api_engagements(request):
    org = request.organization
    qs = Engagement.objects.filter(organization=org).order_by('title')
    data = [{'id': e.pk, 'text': e.title} for e in qs]
    return JsonResponse(data, safe=False)

@login_required
def api_objectives(request):
    org = request.organization
    engagement_id = request.GET.get('engagement')
    qs = Objective.objects.filter(engagement__organization=org)
    if engagement_id:
        qs = qs.filter(engagement_id=engagement_id)
    data = [{'id': o.pk, 'text': o.title} for o in qs]
    return JsonResponse(data, safe=False)

@login_required
def api_issues(request):
    org = request.organization
    engagement_id = request.GET.get('engagement')
    qs = Issue.objects.filter(organization=org)
    if engagement_id:
        qs = qs.filter(engagement_id=engagement_id)
    data = [{'id': i.pk, 'text': i.issue_title} for i in qs]
    return JsonResponse(data, safe=False)

@login_required
def api_recommendations(request):
    org = request.organization
    issue_id = request.GET.get('issue')
    qs = Recommendation.objects.filter(organization=org)
    if issue_id:
        qs = qs.filter(issue_id=issue_id)
    data = [{'id': r.pk, 'text': r.title} for r in qs]
    return JsonResponse(data, safe=False)

@require_GET
def htmx_objective_list(request):
    # Handle unauthenticated users gracefully
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view objectives.</div>")
        return JsonResponse({'html': '', 'success': False, 'message': 'Authentication required'}, status=200)
    
    try:
        org = request.organization
        engagement_id = request.GET.get('engagement')
        qs = Objective.objects.filter(engagement__organization=org)
        engagement = None
        if engagement_id:
            qs = qs.filter(engagement_id=engagement_id)
            engagement = Engagement.objects.filter(pk=engagement_id, organization=org).first()
        html = render_to_string('audit/_objective_list_partial.html', {
            'objectives': qs,
            'engagement': engagement
        })
        return JsonResponse({'html': html, 'success': True})
    except Exception as e:
        logger.error(f"Error in htmx_objective_list: {str(e)}", exc_info=True)
        return JsonResponse({'html': f"<div class='alert alert-danger'>Error loading objectives.</div>", 'success': False}, status=200)

@require_GET
def htmx_procedure_list(request):
    # Handle unauthenticated users gracefully
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view procedures.</div>")
        return JsonResponse({'html': '', 'success': False, 'message': 'Authentication required'}, status=200)
    
    try:
        org = request.organization
        objective_id = request.GET.get('objective')
        qs = Procedure.objects.filter(objective__engagement__organization=org)
        if objective_id:
            qs = qs.filter(objective_id=objective_id)
        html = render_to_string('audit/_procedure_list_partial.html', {'procedures': qs})
        return JsonResponse({'html': html, 'success': True})
    except Exception as e:
        logger.error(f"Error in htmx_procedure_list: {str(e)}", exc_info=True)
        return JsonResponse({'html': f"<div class='alert alert-danger'>Error loading procedures.</div>", 'success': False}, status=200)

@require_GET
def htmx_recommendation_list(request):
    # Handle unauthenticated users gracefully
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view recommendations.</div>")
        return JsonResponse({'html': '', 'success': False, 'message': 'Authentication required'}, status=200)
    
    try:
        org = request.organization
        issue_id = request.GET.get('issue')
        qs = Recommendation.objects.filter(organization=org)
        if issue_id:
            qs = qs.filter(issue_id=issue_id)
        html = render_to_string('audit/_recommendation_list_partial.html', {'recommendations': qs})
        return JsonResponse({'html': html, 'success': True})
    except Exception as e:
        logger.error(f"Error in htmx_recommendation_list: {str(e)}", exc_info=True)
        return JsonResponse({'html': f"<div class='alert alert-danger'>Error loading recommendations.</div>", 'success': False}, status=200)

@login_required
@require_http_methods(['POST'])
def submit_workplan(request, pk):
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view follow-up actions.</div>")
        return JsonResponse({'success': False, 'message': 'Authentication required'}, status=200)
    try:
        org = request.organization
        workplan = AuditWorkplan.objects.get(pk=pk, organization=org)
        workplan.status = AuditWorkplan.STATUS_APPROVED
        workplan.save()
        return JsonResponse({'success': True, 'message': 'Workplan submitted successfully'})
    except Exception as e:
        logger.error(f"Error in submit_workplan: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'Error submitting workplan'}, status=500)

@login_required
@require_http_methods(['POST'])
def approve_workplan(request, pk):
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view follow-up actions.</div>")
        return JsonResponse({'success': False, 'message': 'Authentication required'}, status=200)
    try:
        org = request.organization
        workplan = AuditWorkplan.objects.get(pk=pk, organization=org)
        workplan.status = AuditWorkplan.STATUS_APPROVED
        workplan.save()
        return JsonResponse({'success': True, 'message': 'Workplan approved successfully'})
    except Exception as e:
        logger.error(f"Error in approve_workplan: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'Error approving workplan'}, status=500)

@login_required
@require_http_methods(['POST'])
def reject_workplan(request, pk):
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view follow-up actions.</div>")
        return JsonResponse({'success': False, 'message': 'Authentication required'}, status=200)
    try:
        org = request.organization
        workplan = AuditWorkplan.objects.get(pk=pk, organization=org)
        workplan.status = AuditWorkplan.STATUS_REJECTED
        workplan.save()
        return JsonResponse({'success': True, 'message': 'Workplan rejected successfully'})
    except Exception as e:
        logger.error(f"Error in reject_workplan: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'Error rejecting workplan'}, status=500)

@login_required
@require_http_methods(['POST'])
def submit_followupaction(request, pk):
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view follow-up actions.</div>")
        return JsonResponse({'success': False, 'message': 'Authentication required'}, status=200)
    try:
        org = request.organization
        followupaction = FollowUpAction.objects.get(pk=pk, organization=org)
        followupaction.status = FollowUpAction.STATUS_APPROVED
        followupaction.save()
        return JsonResponse({'success': True, 'message': 'Follow-up action submitted successfully'})
    except Exception as e:
        logger.error(f"Error in submit_followupaction: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'Error submitting follow-up action'}, status=500)

@login_required
@require_http_methods(['POST'])
def approve_followupaction(request, pk):
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view follow-up actions.</div>")
        return JsonResponse({'success': False, 'message': 'Authentication required'}, status=200)
    try:
        org = request.organization
        followupaction = FollowUpAction.objects.get(pk=pk, organization=org)
        followupaction.status = FollowUpAction.STATUS_APPROVED
        followupaction.save()
        return JsonResponse({'success': True, 'message': 'Follow-up action approved successfully'})
    except Exception as e:
        logger.error(f"Error in approve_followupaction: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'Error approving follow-up action'}, status=500)

@login_required
@require_http_methods(['POST'])
def reject_followupaction(request, pk):
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view follow-up actions.</div>")
        return JsonResponse({'success': False, 'message': 'Authentication required'}, status=200)
    try:
        org = request.organization
        followupaction = FollowUpAction.objects.get(pk=pk, organization=org)
        followupaction.status = FollowUpAction.STATUS_REJECTED
        followupaction.save()
        return JsonResponse({'success': True, 'message': 'Follow-up action rejected successfully'})
    except Exception as e:
        logger.error(f"Error in reject_followupaction: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'Error rejecting follow-up action'}, status=500)

@login_required
@require_http_methods(['GET'])
def htmx_followupaction_list(request):
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view follow-up actions.</div>")
        return JsonResponse({'success': False, 'message': 'Authentication required'}, status=200)
    try:
        # Get organization from the request or user's primary organization
        org = getattr(request, 'organization', None)
        if not org and hasattr(request.user, 'organization'):
            org = request.user.organization
            
        if not org:
            logger.error("Organization not found in request or user profile")
            raise PermissionDenied("Organization not found. Please ensure you're accessing this from within an organization context.")

        # Get filter parameters
        issue_id = request.GET.get('issue_id')
        recommendation_id = request.GET.get('recommendation')
        
        # Base queryset with select_related for performance
        qs = FollowUpAction.objects.filter(organization=org)
        qs = qs.select_related('issue', 'assigned_to', 'created_by')
        
        # Apply filters
        if issue_id and issue_id.isdigit():
            qs = qs.filter(issue_id=int(issue_id))
        if recommendation_id and recommendation_id.isdigit():
            qs = qs.filter(recommendation_id=int(recommendation_id))
        
        # Handle ordering with nulls last for due_date
        try:
            # Try with nulls_last if the database supports it
            qs = qs.order_by(F('due_date').asc(nulls_last=True), '-created_at')
        except (NotSupportedError, DatabaseError):
            # Fallback for databases that don't support nulls_last
            qs = qs.order_by(
                Case(When(due_date__isnull=True, then=1), default=0),
                     'due_date',
                     '-created_at'
            )
        
        context = {
            'followups': qs,
            'issue_id': issue_id,
            'request': request
        }
        
        # If this is an HTMX request, return just the list content
        if request.headers.get('HX-Request'):
            html = render_to_string('audit/_followupaction_list_partial.html', context)
            return HttpResponse(html)
        
        # For regular AJAX requests, return JSON
        return JsonResponse({
            'success': True,
            'count': qs.count(),
            'html': render_to_string('audit/_followupaction_list_partial.html', context)
        })
        
    except Exception as e:
        error_msg = f"Error in htmx_followupaction_list: {str(e)}"
        logger.error(error_msg, exc_info=True)
        
        # Prepare error context
        error_context = {
            'error': str(e),
            'error_type': e.__class__.__name__,
            'issue_id': request.GET.get('issue_id'),
            'recommendation_id': request.GET.get('recommendation'),
            'user': request.user.username if request.user.is_authenticated else 'anonymous'
        }
        
        # Log detailed error context
        logger.error(f"Error context: {error_context}")
        
        # Return appropriate response based on request type
        if request.headers.get('HX-Request'):
            error_html = render_to_string('audit/_error_alert.html', {
                'title': 'Error Loading Follow-up Actions',
                'message': 'There was an error loading the follow-up actions. Please try again later.',
                'error': str(e),
                'error_type': e.__class__.__name__,
                'debug': getattr(settings, 'DEBUG', False)
            })
            return HttpResponse(error_html, status=500)
            
        return JsonResponse(
            {
                'success': False,
                'error': 'An error occurred while loading follow-up actions.',
                'detail': str(e),
                'error_type': e.__class__.__name__
            },
            status=500
        )

@require_GET
def htmx_issueretest_list(request):
    # Handle unauthenticated users gracefully
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view issue retests.</div>")
        return JsonResponse({'html': '', 'success': False, 'message': 'Authentication required'}, status=200)
    
    try:
        org = request.organization
        recommendation_id = request.GET.get('recommendation')
        qs = IssueRetest.objects.filter(organization=org)
        if recommendation_id:
            qs = qs.filter(recommendation_id=recommendation_id)
        html = render_to_string('audit/_issueretest_list_partial.html', {'issueretests': qs})
        return JsonResponse({'html': html, 'success': True})
    except Exception as e:
        logger.error(f"Error in htmx_issueretest_list: {str(e)}", exc_info=True)
        return JsonResponse({'html': f"<div class='alert alert-danger'>Error loading issue retests.</div>", 'success': False}, status=200)

@require_GET
def htmx_note_list(request):
    # Handle unauthenticated users gracefully
    if not request.user.is_authenticated:
        if request.headers.get('HX-Request'):
            return HttpResponse("<div class='alert alert-info'>Please log in to view notes.</div>")
        return JsonResponse({'html': '', 'success': False, 'message': 'Authentication required'}, status=200)
    
    try:
        org = request.organization
        object_id = request.GET.get('object_id')
        content_type_id = request.GET.get('content_type_id')
        qs = Note.objects.filter(organization=org)
        if object_id and content_type_id:
            qs = qs.filter(object_id=object_id, content_type_id=content_type_id)
        html = render_to_string('audit/_note_list_partial.html', {'notes': qs})
        return JsonResponse({'html': html, 'success': True})
    except Exception as e:
        logger.error(f"Error in htmx_note_list: {str(e)}", exc_info=True)
        return JsonResponse({'html': f"<div class='alert alert-danger'>Error loading notes.</div>", 'success': False}, status=200)

class NoteDetailView(AuditPermissionMixin, DetailView):
    model = Note
    template_name = 'audit/note_detail.html'
    context_object_name = 'note'

# ─── ISSUE WORKING PAPER VIEWS ───────────────────────────────────────────────
class IssueWorkingPaperListView(AuditPermissionMixin, ListView):
    model = IssueWorkingPaper
    template_name = 'audit/issueworkingpaper_list.html'
    context_object_name = 'working_papers'
    paginate_by = 20

    def get_queryset(self):
        issue_pk = self.kwargs.get('issue_pk')
        organization = self.request.organization
        return IssueWorkingPaper.objects.filter(issue_id=issue_pk, organization=organization)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['issue_pk'] = self.kwargs.get('issue_pk')
        return context

class IssueWorkingPaperCreateView(AuditPermissionMixin, SuccessMessageMixin, CreateView):
    model = IssueWorkingPaper
    form_class = IssueWorkingPaperForm
    template_name = 'audit/issueworkingpaper_form.html'
    success_message = _('Working paper uploaded successfully')

    def get_template_names(self):
        if self.request.htmx or self.request.headers.get('HX-Request'):
            return ['audit/issueworkingpaper_modal_form.html']
        return [self.template_name]

    def form_valid(self, form):
        form.instance.organization = self.request.organization  # Robust fix for IntegrityError
        response = super().form_valid(form)
        if self.request.htmx or self.request.headers.get('HX-Request'):
            working_papers = IssueWorkingPaper.objects.filter(issue=self.object.issue)
            html_list = render_to_string('audit/issueworkingpaper_list.html', {
                'working_papers': working_papers,
                'issue': self.object.issue,
            }, request=self.request)
            return JsonResponse({'form_is_valid': True, 'html_list': html_list})
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if not hasattr(self, 'object') or not self.object:
            context['issue'] = get_object_or_404(Issue, pk=self.kwargs.get('issue_pk'))
        return context

    def form_invalid(self, form):
        if self.request.htmx or self.request.headers.get('HX-Request'):
            html_form = render_to_string('audit/issueworkingpaper_modal_form.html', 
                self.get_context_data(form=form, object=None),
                request=self.request
            )
            return JsonResponse({'form_is_valid': False, 'html_form': html_form})
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('audit:issue-detail', kwargs={'pk': self.object.issue.pk})

class IssueWorkingPaperUpdateView(AuditPermissionMixin, SuccessMessageMixin, UpdateView):
    model = IssueWorkingPaper
    form_class = IssueWorkingPaperForm
    template_name = 'audit/issueworkingpaper_form.html'
    success_message = _('Working paper updated successfully')

    def get_template_names(self):
        if self.request.htmx or self.request.headers.get('HX-Request'):
            return ['audit/issueworkingpaper_modal_form.html']
        return [self.template_name]

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.htmx or self.request.headers.get('HX-Request'):
            working_papers = IssueWorkingPaper.objects.filter(issue=self.object.issue)
            html_list = render_to_string('audit/issueworkingpaper_list.html', {
                'working_papers': working_papers,
                'issue': self.object.issue,
            }, request=self.request)
            return JsonResponse({'form_is_valid': True, 'html_list': html_list})
        return response

    def form_invalid(self, form):
        if self.request.htmx or self.request.headers.get('HX-Request'):
            html_form = render_to_string('audit/issueworkingpaper_modal_form.html', {
                'form': form,
                'object': self.object,
            }, request=self.request)
            return JsonResponse({'form_is_valid': False, 'html_form': html_form})
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('audit:issue-detail', kwargs={'pk': self.object.issue_id})

class IssueWorkingPaperDeleteView(AuditPermissionMixin, DeleteView):
    model = IssueWorkingPaper
    template_name = 'audit/issueworkingpaper_confirm_delete.html'

    def get_template_names(self):
        if self.request.htmx or self.request.headers.get('HX-Request'):
            return ['audit/issueworkingpaper_confirm_delete.html']
        return [self.template_name]

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        issue = self.object.issue
        self.object.delete()
        if self.request.htmx or self.request.headers.get('HX-Request'):
            working_papers = IssueWorkingPaper.objects.filter(issue=issue)
            html_list = render_to_string('audit/issueworkingpaper_list.html', {
                'working_papers': working_papers,
                'issue': issue,
            }, request=self.request)
            return JsonResponse({'form_is_valid': True, 'html_list': html_list})
        return super().delete(request, *args, **kwargs)

    def get_form_kwargs(self):
        # Do NOT pass 'organization' or any extra kwargs
        return super().get_form_kwargs()

class IssueWorkingPaperDetailView(AuditPermissionMixin, DetailView):
    model = IssueWorkingPaper
    template_name = 'audit/issueworkingpaper_detail.html'
    context_object_name = 'working_paper'

# ─── API VIEWSET ─────────────────────────────────────────────────────────────
from rest_framework import mixins, viewsets
from rest_framework.response import Response
from rest_framework import status
from .serializers import (RiskSerializer, RiskDetailSerializer, 
                         IssueSerializer, ApprovalSerializer,
                         IssueWorkingPaperSerializer)
class IssueWorkingPaperViewSet(OrganizationScopedApiMixin, viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated, IsOrgManagerOrReadOnly]
    
    def get_serializer_class(self):
        from .serializers import IssueWorkingPaperSerializer
        return IssueWorkingPaperSerializer

    def get_queryset(self):
        # Get base queryset with organization filtering from mixin
        qs = super().get_queryset()
        
        # Apply additional filters
        issue_pk = self.kwargs.get('issue_pk')
        if issue_pk:
            qs = qs.filter(issue_id=issue_pk)
            
        return qs
        
# Risk API ViewSet        
class RiskViewSet(OrganizationScopedApiMixin, viewsets.ModelViewSet):
    """API ViewSet for managing Risk objects.
    Provides CRUD operations with proper organization scoping and permission checks.
    Uses OrganizationScopedApiMixin for consistent organization filtering.
    """
    permission_classes = [permissions.IsAuthenticated, IsOrgManagerOrReadOnly]
    
    def get_serializer_class(self):
        from .serializers import RiskSerializer, RiskDetailSerializer
        if self.action == 'retrieve' or self.action == 'update' or self.action == 'partial_update':
            return RiskDetailSerializer
        if self.action == 'list':
            # Use RiskSerializer for list view
            return RiskSerializer
        return RiskSerializer
    
    def get_queryset(self):
        # Get base queryset with organization filtering from mixin
        qs = super().get_queryset()
        
        # Apply additional filters
        objective_pk = self.request.query_params.get('objective_id')
        engagement_pk = self.request.query_params.get('engagement_id')
        
        # Apply filters if provided
        if objective_pk:
            qs = qs.filter(objective_id=objective_pk)
        if engagement_pk:
            qs = qs.filter(objective__engagement_id=engagement_pk)
            
        # Apply status filter if provided
        status = self.request.query_params.get('status')
        if status:
            qs = qs.filter(status=status)
            
        # Apply category filter if provided
        category = self.request.query_params.get('category')
        if category:
            qs = qs.filter(category=category)
            
        # Order risks by order field, then by title
        return qs.order_by('order', 'title')


class ObjectiveViewSet(OrganizationScopedApiMixin, viewsets.ModelViewSet):
    """API ViewSet for managing Objective objects.
    Provides CRUD operations with proper organization scoping and permission checks.
    Uses OrganizationScopedApiMixin for consistent organization filtering.
    """
    permission_classes = [permissions.IsAuthenticated, IsOrgManagerOrReadOnly]
    
    def get_serializer_class(self):
        from .serializers import ObjectiveSerializer
        if self.action in ['retrieve', 'update', 'partial_update']:
            # Use ObjectiveDetailSerializer if available
            try:
                from .serializers import ObjectiveDetailSerializer
                return ObjectiveDetailSerializer
            except ImportError:
                return ObjectiveSerializer
        return ObjectiveSerializer
    
    def get_queryset(self):
        from .models.objective import Objective
        # Get base queryset with organization filtering from mixin
        qs = super().get_queryset()
        
        # Apply additional filters
        engagement_pk = self.request.query_params.get('engagement_id')
        
        # Apply filters if provided
        if engagement_pk:
            qs = qs.filter(engagement_id=engagement_pk)
        
        # Order objectives by order field, then by title
        return qs.order_by('order', 'title')
    
    def perform_create(self, serializer):
        serializer.save(
            organization=self.request.organization,
            created_by=self.request.user
        )

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required

@login_required
def api_issue_risk_data(request):
    org = request.user.organization
    from .models.issue import Issue
    from django.db.models import Count
    risk_qs = Issue.objects.filter(organization=org).values('risk_level').annotate(count=Count('id'))
    risk_data = {s['risk_level']: s['count'] for s in risk_qs}
    return JsonResponse(risk_data, safe=False)

@login_required
def api_approval_status_data(request):
    org = request.user.organization
    from .models.approval import Approval
    from django.db.models import Count
    status_qs = Approval.objects.filter(organization=org).values('status').annotate(count=Count('id'))
    status_data = {s['status']: s['count'] for s in status_qs}
    return JsonResponse(status_data, safe=False)

@login_required
def api_engagement_status_data(request):
    org = request.user.organization
    from .models.engagement import Engagement
    status_qs = Engagement.objects.filter(organization=org).values('project_status').annotate(count=Count('id'))
    status_data = {s['project_status']: s['count'] for s in status_qs}
    return JsonResponse(status_data, safe=False)

@login_required
def api_engagement_data(request):
    org = request.user.organization
    from .models.engagement import Engagement
    from django.db.models import Count
    engagement_qs = Engagement.objects.filter(organization=org).values('project_status').annotate(count=Count('id'))
    engagement_data = {s['project_status']: s['count'] for s in engagement_qs}
    return JsonResponse(engagement_data, safe=False)

@login_required
def api_issue_data(request):
    """API endpoint for issue data used in dashboards."""
    issues = Issue.objects.filter(organization=request.user.organization)
    data = {
        'total': issues.count(),
        'open': issues.filter(issue_status='open').count(),
        'in_progress': issues.filter(issue_status='in_progress').count(),
        'closed': issues.filter(issue_status='closed').count(),
        'high_risk': issues.filter(risk_level='high').count(),
        'medium_risk': issues.filter(risk_level='medium').count(),
        'low_risk': issues.filter(risk_level='low').count(),
    }
    return JsonResponse(data)

class IssueRetestDeleteView(AuditPermissionMixin, DeleteView):
    model = IssueRetest
    template_name = 'audit/issueretest_confirm_delete.html'

    def get_template_names(self):
        if self.request.htmx or self.request.headers.get('HX-Request'):
            return ['audit/issueretest_confirm_delete.html']
        return [self.template_name]

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        issue = self.object.issue
        self.object.delete()
        if self.request.htmx or self.request.headers.get('HX-Request'):
            from django.template.loader import render_to_string
            from django.http import JsonResponse
            retests = IssueRetest.objects.filter(issue=issue)
            html_list = render_to_string('audit/_issueretest_list_partial.html', {
                'retests': retests,
                'issue': issue,
            }, request=self.request)
            return JsonResponse({'form_is_valid': True, 'html_list': html_list})
        return super().delete(request, *args, **kwargs)

    def get_form_kwargs(self):
        return super().get_form_kwargs()

    def get_success_url(self):
        return reverse_lazy('audit:issue-detail', kwargs={'pk': self.object.issue.pk})
